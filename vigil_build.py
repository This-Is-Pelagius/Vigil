#!/usr/bin/env python3
"""
vigil_build.py
Vigil PWA build script — Version 2.0
Reads Vigil_Content_v2.0.xlsx, identifies today's day, and writes a single
index.html for deployment to the v2-build branch of github.com/This-Is-Pelagius/Vigil.

Usage:
    python3 vigil_build.py

Output:
    index.html  (in the same directory as this script)
"""

import re
import sys
import html
import datetime
import os

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install it with: pip3 install openpyxl")
    sys.exit(1)

# ── Configuration ─────────────────────────────────────────────────────────────

SPREADSHEET = "Vigil_Content_v2.0.xlsx"
OUTPUT_FILE = "index.html"

# Number of days to embed ahead of today.
# The app selects today's content at runtime, so users only see today.
# Embedding future days means the midnight reload finds tomorrow already
# in the file, giving automatic time-zone-correct daily advance without
# requiring a new deployment every day.
DAYS_WINDOW = 14

# Liturgical season → CSS palette class mapping
# Used as fallback when the Palette column in the spreadsheet is blank.
PALETTE_MAP = {
    "eastertide":    "palette-eastertide",
    "lent":          "palette-lent",
    "good friday":   "palette-goodfriday",
    "holy saturday": "palette-goodfriday",
    "pentecost":     "palette-pentecost",
    "ordinary time": "palette-ordinarytime",
    "advent":        "palette-advent",
}

# Spreadsheet Palette column value → CSS palette class
# This is the primary palette control. When a value is present in the
# spreadsheet Palette column, it takes precedence over everything else.
SPREADSHEET_PALETTE_MAP = {
    "eastertide":   "palette-eastertide",
    "red major":    "palette-pentecost",
    "red":          "palette-red",
    "green major":  "palette-ordinarytime-major",
    "green":        "palette-ordinarytime",
    "white major":  "palette-white-major",
    "white":        "palette-white",
    "purple major": "palette-lent",
    "purple":       "palette-advent",
    "rose":         "palette-rose",
    "black":        "palette-black",
    "good friday":  "palette-goodfriday",
    "lent":         "palette-lent",
}

# Feast/solemnity names in the feast line that override the season palette.
# Used as a secondary fallback when the Palette column is blank and the
# season key alone is insufficient (e.g. Pentecost within Eastertide).
FEAST_PALETTE_OVERRIDE = {
    "pentecost":   "palette-pentecost",
    "good friday": "palette-goodfriday",
}

# Abbreviated month names for the day label and word echo date
MONTH_ABBR = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}

# Screen labels (indices 0–6)
SCREEN_LABELS = ["Season", "Word", "Scripture", "Contemplation", "Prayer", "Practice", "Amen"]

# Jesus is never hover-linked — any variant must be excluded
JESUS_NAMES = {
    "jesus", "jesus christ", "christ jesus", "lord jesus", "lord jesus christ",
}

# ── Spreadsheet parsing ───────────────────────────────────────────────────────

def load_spreadsheet(path):
    """Load all data rows from all tabs of the spreadsheet.
    Returns a list of dicts, one per content day, in sheet order."""
    if not os.path.exists(path):
        print(f"Error: Cannot find spreadsheet '{path}'.")
        print("Make sure Vigil_Content_v2.0.xlsx is in the same folder as this script.")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    days = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            continue

        # Row 1 is the header
        header = [str(c).strip() if c else "" for c in rows[0]]

        # Build a column index map
        col = {}
        for i, h in enumerate(header):
            hl = h.lower()
            if "version" in hl:
                col["version"] = i
            elif "palette" in hl:
                col["palette"] = i
            elif "liturgical day" in hl:
                col["liturgical_day"] = i
            elif "screen 1" in hl or "season" in hl:
                col["season"] = i
            elif "screen 2" in hl or "word" in hl:
                col["word"] = i
            elif "screen 3" in hl or "scripture" in hl:
                col["scripture"] = i
            elif "screen 4" in hl or "contemplation" in hl:
                col["contemplation"] = i
            elif "screen 5" in hl or "prayer" in hl:
                col["prayer"] = i
            elif "screen 6" in hl or "practice" in hl:
                col["practice"] = i
            elif "hover" in hl:
                col["hover_links"] = i

        # Required columns
        required = ["liturgical_day", "season", "word", "scripture",
                    "contemplation", "prayer", "practice"]
        for r in required:
            if r not in col:
                print(f"Warning: Sheet '{sheet_name}' — could not find column for '{r}'. Skipping sheet.")
                col = {}
                break
        if not col:
            continue

        for row in rows[1:]:
            # Skip empty rows
            if all(v is None or str(v).strip() == "" for v in row):
                continue

            def get(key, default=""):
                if key not in col:
                    return default
                v = row[col[key]]
                if v is None:
                    return default
                s = str(v).strip()
                # Strip leading single quote (Excel text-prefix marker)
                if s.startswith("'"):
                    s = s[1:]
                return s

            day = {
                "version":      get("version"),
                "palette":      get("palette"),
                "liturgical_day": get("liturgical_day"),
                "season":       get("season"),
                "word":         get("word"),
                "scripture":    get("scripture"),
                "contemplation": get("contemplation"),
                "prayer":       get("prayer"),
                "practice":     get("practice"),
                "hover_links":  get("hover_links"),
                "sheet":        sheet_name,
            }

            # Must have at minimum a liturgical day and a word
            if not day["liturgical_day"] or not day["word"]:
                continue

            days.append(day)

    return days


# ── Date parsing ──────────────────────────────────────────────────────────────

MONTH_NAMES = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}

def parse_date_from_liturgical_day(liturgical_day_text):
    """Extract a date from the position line in the Liturgical Day cell.
    The position line contains a date like '25 April 2026'.
    Returns a datetime.date or None."""
    # Try to find a pattern like "25 April 2026" or "1 May 2026"
    m = re.search(r'\b(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})\b', liturgical_day_text)
    if not m:
        return None
    day_n = int(m.group(1))
    month_n = MONTH_NAMES.get(m.group(2).lower())
    year_n = int(m.group(3))
    if not month_n:
        return None
    try:
        return datetime.date(year_n, month_n, day_n)
    except ValueError:
        return None


def today_date():
    return datetime.date.today()


def find_today(days):
    """Return the index of today's day in the days list, or -1."""
    today = today_date()
    for i, day in enumerate(days):
        d = parse_date_from_liturgical_day(day["liturgical_day"])
        if d and d == today:
            return i
    return -1


# ── Liturgical Day parsing ────────────────────────────────────────────────────

# Ordinal word map for normalising numeric ordinals in day names
# e.g. "5th Sunday of Easter" → "Fifth Sunday of Easter"
ORDINAL_WORDS = {
    "1st": "First",  "2nd": "Second", "3rd": "Third",  "4th": "Fourth",
    "5th": "Fifth",  "6th": "Sixth",  "7th": "Seventh","8th": "Eighth",
    "9th": "Ninth",  "10th": "Tenth",
}

def normalise_ordinal(text):
    """Replace numeric ordinals with words: '5th Sunday' → 'Fifth Sunday'."""
    return re.sub(
        r'\b(\d{1,2}(?:st|nd|rd|th))\b',
        lambda m: ORDINAL_WORDS.get(m.group(1).lower(), m.group(1)),
        text,
    )


def parse_liturgical_day(text):
    """Parse the Liturgical Day cell.

    Actual cell structure (confirmed against spreadsheet):
      Line 0: date          e.g. "29 April 2026"
      Line 1: day name      e.g. "Wednesday" / "Fourth Sunday of Easter"
      Line 2: season·week   e.g. "Eastertide · Week Four"
      Line 3: feast line    e.g. "Memorial · St Catherine of Siena…" (optional)

    Returns (season_name, position_line, feast_line, season_key) where:
      season_name   = the season·week string (line 2), used for palette lookup
      position_line = the full original text (legacy — used by format_day_label
                      and parse_date_from_liturgical_day via regex)
      feast_line    = line 3 if present, else ""
      season_key    = first token of season_name, lowercased, e.g. "eastertide"
    """
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    # season_name: take the season·week line (index 2) for palette detection
    season_name   = lines[2] if len(lines) > 2 else (lines[0] if lines else "")
    position_line = text  # keep full text so date regex still works
    feast_line    = lines[3] if len(lines) > 3 else ""
    # season_key: first word of the season·week line, e.g. "eastertide"
    season_key    = season_name.split()[0].lower() if season_name else ""
    return season_name, position_line, feast_line, season_key


def format_day_label(season_name, position_line):
    """Build the top-bar day label: '29 April 2026' (full date, no season name)."""
    m = re.search(r'\b(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})\b', position_line)
    if m:
        return f"{int(m.group(1))} {m.group(2).capitalize()} {m.group(3)}"
    return ""


def format_word_echo_date(season_name, position_line):
    """Build the word echo date line: '29 April 2026' (matches top bar label)."""
    return format_day_label(season_name, position_line)


# ── Hover link parsing ────────────────────────────────────────────────────────

def parse_hover_links(hover_links_text):
    """Parse the Hover Links cell into a dict keyed by figure slug.
    Each block is separated by a blank line.
    Format: Name (dates).\nBody text.
    Pipe format: Role | Name (dates).\nBody text.
    Returns {slug: {name, dates, role, body}}"""
    if not hover_links_text.strip():
        return {}

    figures = {}
    # Split on blank lines
    blocks = re.split(r'\n\s*\n', hover_links_text.strip())

    for block in blocks:
        block = block.strip()
        if not block:
            continue

        lines = [l.strip() for l in block.split("\n") if l.strip()]
        if not lines:
            continue

        header = lines[0]
        body_lines = lines[1:]
        body = " ".join(body_lines).strip()

        role = ""

        # Pipe format: Role | Name (dates).
        # Also accepts middle dot (·, U+00B7) as separator, e.g. "Author · Psalm 130 (dates)."
        pipe_sep = None
        if "|" in header:
            pipe_sep = "|"
        elif "\u00b7" in header:
            pipe_sep = "\u00b7"
        if pipe_sep:
            parts = header.split(pipe_sep, 1)
            role = parts[0].strip()
            header = parts[1].strip()
            # Special case: when role and name represent the same figure
            # (e.g. "Psalmist · Psalm 130"), use the role as the display name
            # and suppress the role line so the card reads cleanly.
            if role.lower() == "psalmist":
                # Preserve dates from the name field (e.g. "Psalm 130 (dates)")
                # but display as "Psalmist" with those dates.
                header = "Psalmist" + (" (" + header.split("(", 1)[1] if "(" in header else "")
                role = ""

        # Extract name and dates: Name (dates). or Name (dates)
        m = re.match(r'^(.+?)\s*\(([^)]+)\)\.?\s*$', header)
        if m:
            name = m.group(1).strip()
            dates = m.group(2).strip()
        else:
            # No dates found — use the whole header as name
            name = header.rstrip(".")
            dates = ""

        slug = make_figure_slug(name)
        figures[slug] = {
            "name":  name,
            "dates": dates,
            "role":  role,
            "body":  body,
        }

    return figures


def make_figure_slug(name):
    """Convert a figure name to a slug key, e.g. 'St Mark, Evangelist' → 'st_mark_evangelist'"""
    s = name.lower()
    s = re.sub(r'[^a-z0-9\s]', '', s)
    s = re.sub(r'\s+', '_', s.strip())
    return s


# ── Figure link injection ─────────────────────────────────────────────────────

def build_figure_link_pattern(figures):
    """Build a regex that matches any figure name in the figures dict.
    Returns (pattern, slug_map) or (None, {}) if no figures."""
    if not figures:
        return None, {}

    # Build a map from (lower-case name) → slug
    name_to_slug = {}
    for slug, fig in figures.items():
        name_lower = fig["name"].lower()
        if name_lower not in JESUS_NAMES:
            name_to_slug[name_lower] = slug

    if not name_to_slug:
        return None, {}

    # Also add short-name variants: "St Mark, Evangelist" → also match "Mark"
    # For joint entries like "Sts Nereus and Achilleus", register every proper
    # name token so both names are hover-linked in the body copy.
    SKIP_WORDS = {'and', 'of', 'the', 'sts', 'st', 'saint', 'blessed', 'fr',
                  'dr', 'pope', 'bishop'}
    extra = {}
    for name_lower, slug in list(name_to_slug.items()):
        # Strip honorifics and role suffixes
        clean = re.sub(r'\b(st|sts|saint|blessed|fr|dr|pope|bishop)\b', '', name_lower)
        clean = re.sub(r',.*$', '', clean)  # remove role suffix after comma
        clean = clean.strip()
        tokens = clean.split()
        # Register every proper-name token (length ≥ 3, not a connective word)
        for token in tokens:
            if (len(token) >= 3
                    and token not in SKIP_WORDS
                    and token not in JESUS_NAMES
                    and token not in name_to_slug):
                extra[token] = slug

    name_to_slug.update(extra)

    # Psalmist alias: a figure named "Psalmist" should be linked on the word
    # "Psalmist" wherever it appears in the body copy. "Psalm" is not linked
    # as it is too common a word to link safely across all content.
    for slug, fig in figures.items():
        name_l = fig["name"].lower()
        if name_l == "psalmist" or re.match(r'^psalm\s+\d', name_l):
            if "psalmist" not in name_to_slug:
                name_to_slug["psalmist"] = slug

    # Sort by length descending so longer names match before shorter substrings
    sorted_names = sorted(name_to_slug.keys(), key=len, reverse=True)

    # Build pattern — use word boundaries
    escaped = [re.escape(n) for n in sorted_names]
    pattern = re.compile(r'\b(' + '|'.join(escaped) + r')\b', re.IGNORECASE)

    return pattern, name_to_slug


def inject_figure_links(text, pattern, name_to_slug, no_links=False):
    """Replace named figures in text with hover-link button elements.
    If no_links is True (e.g. Screen 5 Prayer), returns plain text."""
    if no_links or pattern is None or not text:
        return html.escape(text)

    result = []
    last = 0
    for m in pattern.finditer(text):
        start, end = m.start(), m.end()
        matched = m.group(1)
        slug = name_to_slug.get(matched.lower())
        if not slug:
            result.append(html.escape(text[last:end]))
            last = end
            continue
        result.append(html.escape(text[last:start]))
        result.append(
            f'<button type="button" class="figure-link" '
            f'data-figure="{slug}">{html.escape(matched)}</button>'
        )
        last = end
    result.append(html.escape(text[last:]))
    return "".join(result)


# ── Screen content builders ───────────────────────────────────────────────────

def paragraphs_to_html(text, pattern=None, name_to_slug=None, no_links=False):
    """Split text on blank lines and return <p> elements.
    Named figures are linked unless no_links is True."""
    paras = [p.strip() for p in re.split(r'\n\s*\n', text.strip()) if p.strip()]
    out = []
    for p in paras:
        inner = inject_figure_links(p, pattern, name_to_slug or {}, no_links=no_links)
        out.append(f"<p>{inner}</p>")
    return "\n".join(out)


def parse_screen1_lines(text):
    """Extract (day_name, season_week) from the Liturgical Day cell.

    Cell structure:
      Line 0: date          "29 April 2026"
      Line 1: day name      "Wednesday" / "Fourth Sunday of Easter"
      Line 2: season·week   "Eastertide · Week Four"
      Line 3: feast line    optional

    Returns (day_name, season_week).
    day_name has numeric ordinals normalised: "5th" → "Fifth".
    season_week is absent ("") for superseding Solemnities that have no
    week descriptor — not currently in the dataset but handled defensively.
    """
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    day_name    = normalise_ordinal(lines[1]) if len(lines) > 1 else ""
    season_week = lines[2]                    if len(lines) > 2 else ""
    # A superseding Solemnity would have no season_week line — leave as "".
    return day_name, season_week


def build_screen1_season(day, pattern, name_to_slug):
    """Build the inner HTML for Screen 1 — Season.

    Header (new design, reading order):
      .season-day    — always   e.g. "Wednesday" / "Fourth Sunday of Easter"
      .season-week   — cond.    e.g. "Eastertide · Week Four" (absent for superseding Solemnities)
      .season-feast  — cond.    e.g. "Memorial · St Catherine of Siena…"
    """
    _, position_line, feast_line, _ = parse_liturgical_day(day["liturgical_day"])
    body_text = day["season"]

    day_name, season_week = parse_screen1_lines(day["liturgical_day"])

    day_name_html    = html.escape(day_name)
    season_week_html = html.escape(season_week) if season_week else ""

    feast_html = ""
    if feast_line:
        feast_inner = inject_figure_links(feast_line, pattern, name_to_slug)
        feast_html = f'<div class="season-feast">{feast_inner}</div>'

    divider_html = '<div class="season-divider"></div>'
    body_paras   = paragraphs_to_html(body_text, pattern, name_to_slug)
    body_html    = f'<div class="season-body">{body_paras}</div>'

    header = f'<div class="season-day">{day_name_html}</div>\n'
    if season_week_html:
        header += f'<div class="season-week">{season_week_html}</div>\n'
    if feast_html:
        header += feast_html + "\n"

    return header + divider_html + "\n" + body_html


def build_screen2_word(day, pattern, name_to_slug):
    """Build the inner HTML for Screen 2 — Word and Definition."""
    raw = day["word"].strip()
    lines = [l.strip() for l in raw.split("\n") if l.strip()]

    # Line 1: WORD or WORD [pronunciation] on same line
    # Line 2 (if starts with [): pronunciation
    # Remaining: definition paragraphs
    word_name = ""
    pronunciation = ""
    definition_lines = []

    if not lines:
        return ""

    first_line = lines[0]

    # Check if pronunciation is embedded on first line: "ABIDE [uh-byd]"
    m = re.match(r'^([A-Z\s\-\']+?)\s*(\[.+?\])\s*$', first_line)
    if m:
        word_name     = m.group(1).strip()
        pronunciation = m.group(2).strip()
        def_start     = 1
    else:
        word_name = first_line
        def_start = 1
        # Check if next non-blank line is a pronunciation
        if len(lines) > 1 and lines[1].startswith("["):
            pronunciation = lines[1].strip()
            def_start = 2

    # Everything after word + pronunciation is definition
    def_text = "\n\n".join(lines[def_start:])
    # Re-join into proper paragraphs (blank-line-separated)
    raw_after = "\n".join(lines[def_start:])
    # Use the original raw to preserve blank lines
    # Find where definition starts in the raw text
    if pronunciation:
        pron_pos = raw.find(pronunciation)
        if pron_pos != -1:
            def_text = raw[pron_pos + len(pronunciation):].strip()
        else:
            def_text = "\n\n".join(lines[def_start:])
    else:
        word_name_pos = raw.find(first_line)
        def_text = raw[word_name_pos + len(first_line):].strip()

    word_name_html    = html.escape(word_name)
    pronunciation_html = html.escape(pronunciation) if pronunciation else ""
    def_paras = paragraphs_to_html(def_text, pattern, name_to_slug)

    pron_block = ""
    if pronunciation_html:
        pron_block = f'<div class="word-pronunciation">{pronunciation_html}</div>'

    return (
        f'<div class="word-title">{word_name_html}</div>\n'
        + (pron_block + "\n" if pron_block else "")
        + '<div class="word-divider"></div>\n'
        + f'<div class="word-definition">{def_paras}</div>'
    )


def build_screen3_scripture(day, pattern, name_to_slug):
    """Build the inner HTML for Screen 3 — Scripture."""
    raw = day["scripture"].strip()
    lines_raw = raw.split("\n")

    # Split into blank-line-separated blocks
    blocks = []
    current = []
    for line in lines_raw:
        if line.strip() == "":
            if current:
                blocks.append("\n".join(current).strip())
                current = []
        else:
            current.append(line)
    if current:
        blocks.append("\n".join(current).strip())

    if not blocks:
        return ""

    # Identify the reference line: the last block matching a citation pattern
    # Pattern: starts with a book name or number+book, contains chapter:verse
    citation_pattern = re.compile(
        r'^(\d\s+)?[A-Z][a-z].*\d+:\d+|^Psalm\s+\d+',
        re.IGNORECASE
    )

    reference = ""
    verse_blocks = []

    # Check last block first
    if citation_pattern.match(blocks[-1]):
        reference = blocks[-1]
        verse_blocks = blocks[:-1]
    else:
        verse_blocks = blocks

    # Build verse HTML — all verses share one .scripture-block
    verse_parts = []
    for i, vb in enumerate(verse_blocks):
        if i > 0:
            verse_parts.append('<div class="scripture-inter-verse"></div>')
        inner = inject_figure_links(vb, pattern, name_to_slug)
        verse_parts.append(f"<p>{inner}</p>")

    verse_html = "\n".join(verse_parts)
    ref_html   = html.escape(reference) if reference else ""

    scripture_block = (
        f'<div class="scripture-block">\n'
        f'  <div class="scripture-text">{verse_html}</div>\n'
        + (f'</div>\n<div class="scripture-translation">{ref_html}</div>' if ref_html
           else '</div>')
    )

    return f'<div class="scripture-passage">\n{scripture_block}\n</div>'


def build_screen4_contemplation(day, pattern, name_to_slug):
    """Build the inner HTML for Screen 4 — Contemplation."""
    raw = day["contemplation"].strip()
    # Split into blank-line-separated blocks
    blocks = [b.strip() for b in re.split(r'\n\s*\n', raw) if b.strip()]

    if not blocks:
        return ""

    # The closing question is the last block.
    # Detect it: it ends with a "?" or is a short single sentence ending with "?"
    # Also acceptable: two closely joined questions.
    question_block = blocks[-1]
    body_blocks    = blocks[:-1]

    body_html = ""
    if body_blocks:
        paras = []
        for b in body_blocks:
            inner = inject_figure_links(b, pattern, name_to_slug)
            paras.append(f"<p>{inner}</p>")
        body_html = (
            f'<div class="contemplation-body">'
            + "\n".join(paras)
            + '</div>'
        )

    q_inner = inject_figure_links(question_block, pattern, name_to_slug)
    question_html = f'<div class="contemplation-question">{q_inner}</div>'

    return (body_html + "\n" if body_html else "") + question_html


def build_screen5_prayer(day):
    """Build the inner HTML for Screen 5 — Prayer.
    No figure links in the prayer."""
    raw = day["prayer"].strip()
    # Split into blank-line-separated blocks = prayer lines
    lines = [b.strip() for b in re.split(r'\n\s*\n', raw) if b.strip()]

    if not lines:
        return ""

    # The closing line begins "In Christ's name" or "Amen"
    amen_patterns = re.compile(r'^(In Christ|Amen)', re.IGNORECASE)
    prayer_lines  = []
    amen_line     = ""

    for line in lines:
        if amen_patterns.match(line) and not amen_line:
            amen_line = line
        else:
            prayer_lines.append(line)

    # If no explicit amen detected, treat the last line as amen
    if not amen_line and prayer_lines:
        amen_line    = prayer_lines[-1]
        prayer_lines = prayer_lines[:-1]

    prayer_paras = "".join(f"<p>{html.escape(l)}</p>\n" for l in prayer_lines)
    amen_html    = f'<div class="prayer-amen">{html.escape(amen_line)}</div>' if amen_line else ""

    return (
        f'<div class="prayer-block">\n'
        f'  <div class="prayer-body">\n{prayer_paras}  </div>\n'
        + (f"  {amen_html}\n" if amen_html else "")
        + '</div>'
    )


def build_screen6_practice(day, pattern, name_to_slug):
    """Build the inner HTML for Screen 6 — Practice."""
    raw = day["practice"].strip()
    blocks = [b.strip() for b in re.split(r'\n\s*\n', raw) if b.strip()]

    if not blocks:
        return ""

    # The anchor sentence is the last block
    anchor_block = blocks[-1]
    body_blocks  = blocks[:-1]

    body_html = ""
    if body_blocks:
        paras = []
        for b in body_blocks:
            inner = inject_figure_links(b, pattern, name_to_slug)
            paras.append(f"<p>{inner}</p>")
        body_html = (
            '<div class="practice-body">'
            + "\n".join(paras)
            + '</div>'
        )

    anchor_inner = inject_figure_links(anchor_block, pattern, name_to_slug)
    anchor_html  = f'<div class="practice-anchor">{anchor_inner}</div>'

    return (body_html + "\n" if body_html else "") + anchor_html


def build_screen7_amen(day, uid):
    """Build the inner HTML for Screen 7 — Amen (fully auto-generated)."""
    raw_word = day["word"].strip()
    first_line = raw_word.split("\n")[0].strip()

    # Extract just the word name (strip any embedded pronunciation)
    m = re.match(r'^([A-Z\s\-\']+?)\s*\[', first_line)
    if m:
        word_name = m.group(1).strip()
    else:
        # Word name is the first line, capitalise title-case for display
        word_name = first_line.title()

    # Parse date for word echo
    season_name, position_line, _, _ = parse_liturgical_day(day["liturgical_day"])
    echo_date = format_word_echo_date(season_name, position_line)

    w = html.escape(word_name)
    d = html.escape(echo_date)
    u = uid

    return f"""
    <!-- Completion hero -->
    <div class="amen-hero" id="amen-hero-{u}">

      <div class="amen-checkbox-row"
           role="checkbox" aria-checked="false" tabindex="0"
           data-uid="{u}"
           onclick="amenCheck(this)"
           onkeydown="if(event.key===' '||event.key==='Enter'){{event.preventDefault();amenCheck(this)}}">
        <div class="amen-checkbox-outer" id="amen-box-{u}">
          <svg class="amen-tick" width="18" height="18" viewBox="0 0 18 18" fill="none">
            <polyline points="3,9 7,13 15,4.5"
              stroke="currentColor" stroke-width="1.5"
              stroke-linecap="round" stroke-linejoin="round"/>
          </svg>
        </div>
        <div class="amen-checkbox-label">I kept Vigil today.</div>
      </div>

      <div class="amen-word-block" id="amen-word-{u}">
        <div class="amen-word-name">{w}</div>
      </div>

    </div>

    <!-- Commission phrase -->
    <div class="amen-commission" id="amen-commission-{u}">
      <div class="amen-commission-line1">Go in peace.</div>
      <div class="amen-commission-line1">And carry the Word with you.</div>
    </div>

    <!-- Counter block — hidden until COUNT_THRESHOLD reached -->
    <div class="amen-count-block" id="amen-count-{u}">
      <div class="amen-count-number" id="amen-count-num-{u}">—</div>
      <div class="amen-count-label">people have kept vigil since Vigil began.</div>
    </div>

    <!-- Post-liturgy layer: rises 3s after commission animation completes -->
    <div class="amen-post-liturgy" id="amen-post-{u}">

      <!-- Notification prompt — shown once, only if not already subscribed -->
      <!-- State is set by vigilMaybeShowNotifPrompt() based on browser/platform detection -->
      <div class="amen-notif-section" id="amen-notif-{u}">

        <!-- State 1: normal — browser supports web push (Android, desktop) -->
        <div id="amen-notif-normal-{u}" style="display:none">
          <div class="amen-notif-prompt">Would you like Vigil to meet you here again tomorrow?</div>
          <button class="amen-share-btn-primary" onclick="vigilRequestNotification('{u}')">
            Turn on notifications
          </button>
          <button class="amen-not-now-btn" onclick="vigilDismissNotification('{u}')">Not now</button>
        </div>

        <!-- State 2: iOS Safari in browser tab — can install as PWA -->
        <div id="amen-notif-ios-safari-{u}" style="display:none">
          <div class="amen-notif-prompt">Morning notifications are only available when Vigil is set as an app on your iPhone.<br>To do this, tap the Share icon at the bottom of your screen and choose Add to Home Screen.</div>
          <button class="amen-not-now-btn" onclick="vigilDismissNotification('{u}')">Dismiss</button>
        </div>

        <!-- State 3: iOS other browser — must switch to Safari -->
        <div id="amen-notif-ios-other-{u}" style="display:none">
          <div class="amen-notif-prompt">Morning notifications are only available when Vigil is set as an app on your iPhone.<br>To do this, open <a href="https://dailyvigil.app" style="color:inherit">dailyvigil.app</a> in Safari and follow the instructions on the Amen screen.</div>
          <button class="amen-not-now-btn" onclick="vigilDismissNotification('{u}')">Dismiss</button>
        </div>

      </div>

      <!-- Share block — conditional on completion count ≥ 2 -->
      <div class="amen-share-block" id="amen-share-{u}">

        <div id="amen-share-initial-{u}">
          <div class="amen-share-prompt">Has Vigil become part of your day?<br>Share it with someone who might find value in it too.</div>
          <button class="amen-share-btn-primary"
                  onclick="amenShareImage('{u}')">Share today&#8217;s Vigil</button>
          <button class="amen-share-btn-secondary"
                  onclick="amenShareApp('{u}')">Share the app</button>
          <button class="amen-not-now-btn"
                  onclick="amenNotNow('{u}')">Not now</button>
        </div>

        <div class="amen-share-confirmed" id="amen-share-confirmed-{u}"></div>

      </div>

    </div>

    <div class="amen-spacer"></div>"""


# ── HTML assembly ─────────────────────────────────────────────────────────────

def build_screen_html(screen_index, inner_html, screen_class):
    """Wrap screen content in the standard screen shell."""
    label = SCREEN_LABELS[screen_index]
    aria  = label
    return f"""    <section class="screen {screen_class}" data-screen="{screen_index}" aria-label="{aria}">
      <div class="screen-fade-top"></div>
      <div class="screen-label">{label}</div>
      <div class="screen-inner">
{inner_html}</div>
      <div class="screen-fade-bottom"></div>
    </section>"""


SCREEN_CLASSES = [
    "screen-season",
    "screen-word",
    "screen-scripture",
    "screen-contemplation",
    "screen-prayer",
    "screen-practice",
    "screen-amen",
]


def build_day_html(day, day_index, figures):
    """Build the HTML for all 7 screens of a single day."""
    pattern, name_to_slug = build_figure_link_pattern(figures)
    uid = f"d{day_index + 1}"

    screen_inners = [
        build_screen1_season(day, pattern, name_to_slug),
        build_screen2_word(day, pattern, name_to_slug),
        build_screen3_scripture(day, pattern, name_to_slug),
        build_screen4_contemplation(day, pattern, name_to_slug),
        build_screen5_prayer(day),
        build_screen6_practice(day, pattern, name_to_slug),
        build_screen7_amen(day, uid),
    ]

    screens_html = []
    for i, inner in enumerate(screen_inners):
        screens_html.append(build_screen_html(i, inner, SCREEN_CLASSES[i]))

    all_screens = "\n".join(screens_html)
    return f'  <div class="day-screens" data-day="{day_index}">\n{all_screens}\n  </div>'


def build_all_days_html(days_window):
    """Build the HTML for all days in the window.
    Each day-screens block is hidden by default; the JS shows the correct one."""
    blocks = []
    for i, (day, figures) in enumerate(days_window):
        block = build_day_html(day, i, figures)
        # All blocks hidden at load time — JS calls showDay() to reveal the right one
        block = block.replace(
            f'data-day="{i}">',
            f'data-day="{i}" style="display:none">',
            1
        )
        blocks.append(block)
    return "\n".join(blocks)


def build_days_js(days_window):
    """Build the DAYS and FIGURES_BY_DAY JavaScript data objects for all days in the window.

    FIGURES_BY_DAY is an array parallel to DAYS — each entry is a dict of figures
    for that specific day only.  This prevents a figure name shared across multiple
    days (e.g. 'John') from having later days' content overwrite earlier days'.
    """
    import json

    days_obj       = []
    figures_by_day = []

    for day, figures in days_window:
        season_name, position_line, feast_line, season_key = parse_liturgical_day(day["liturgical_day"])

        # Priority 1 — Palette column in spreadsheet (explicit instruction)
        spreadsheet_palette = day.get("palette", "").strip().lower()
        if spreadsheet_palette and spreadsheet_palette in SPREADSHEET_PALETTE_MAP:
            palette = SPREADSHEET_PALETTE_MAP[spreadsheet_palette]
        else:
            # Priority 2 — Feast line override (e.g. Pentecost within Eastertide)
            palette = PALETTE_MAP.get(season_key, "palette-ordinarytime")
            feast_lower = feast_line.lower()
            for feast_key, feast_palette in FEAST_PALETTE_OVERRIDE.items():
                if feast_key in feast_lower:
                    palette = feast_palette
                    break
        label    = format_day_label(season_name, position_line)
        d        = parse_date_from_liturgical_day(day["liturgical_day"])
        date_str = d.strftime("%-d %B %Y") if d else ""
        days_obj.append({"palette": palette, "label": label, "date": date_str})
        day_figs = {}
        for slug, fig in figures.items():
            day_figs[slug] = {
                "name":  fig["name"],
                "dates": fig["dates"],
                "role":  fig["role"],
                "body":  fig["body"],
            }
        figures_by_day.append(day_figs)

    days_js    = f"const DAYS           = {json.dumps(days_obj, ensure_ascii=False)};"
    figures_js = f"const FIGURES_BY_DAY = {json.dumps(figures_by_day, ensure_ascii=False, indent=2)};"

    return days_js, figures_js


# ── Full page assembly ────────────────────────────────────────────────────────

CSS = r"""
/* ── Font-loading veil ── */
/* Keeps the app invisible until fonts are ready, preventing the flash of
   fallback text on first load. The JS below lifts this within 800ms. */
.app { opacity: 0; transition: opacity 0.3s ease; }
.app.fonts-ready { opacity: 1; }

*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

:root {
  --bg:          #FAF7F2;
  --panel:       #F2EDE3;
  --accent:      #B8860B;
  --accent-soft: rgba(184,134,11,0.2);
  --text:        #1C1714;
  --dim:         #8A7D6E;
  --gold:        #B8860B;
  --shell:       #F0EAE0;
}

html { font-size: 16px; overflow: hidden; }
html, body {
  height: 100%; margin: 0;
  background: var(--shell);
  font-family: 'Raleway', sans-serif;
  color: var(--text);
  overscroll-behavior: none;
  overflow: hidden;
}
body { overflow: hidden; }

/* ── Palettes ── */
.palette-lent {
  --bg: #241E1A; --panel: #2E2520; --accent: #C49A50; --gold: #C49A50;
  --text: #EDE0C8; --dim: #9A8A78; --shell: #1A1410;
  --accent-soft: rgba(196,154,80,0.2);
}
.palette-goodfriday {
  --bg: #0E0C0B; --panel: #181412; --accent: #A03030; --gold: #A03030;
  --text: #DDD0C0; --dim: #7A6A60; --shell: #080605;
  --accent-soft: rgba(160,48,48,0.2);
}
.palette-pentecost {
  --bg: #8B1A1A; --panel: #7A1515; --accent: #F5D78E; --gold: #F5D78E;
  --text: #FDF6EE; --dim: #E8B89A; --shell: #6B1010;
  --accent-soft: rgba(245,215,142,0.22);
}

/* Ordinary Time — cream background, green accents */
.palette-ordinarytime {
  --bg: #FAF7F2; --panel: #F2EDE3; --accent: #4A7C59; --gold: #4A7C59;
  --text: #1C1714; --dim: #8A7D6E; --shell: #F0EAE0;
  --accent-soft: rgba(74,124,89,0.15);
}

/* Ordinary Time Major — dark green background, gold accents (green solemnities) */
.palette-ordinarytime-major {
  --bg: #1A2B1A; --panel: #162414; --accent: #C8A84B; --gold: #C8A84B;
  --text: #EDE8DC; --dim: #8A9E82; --shell: #111E11;
  --accent-soft: rgba(200,168,75,0.2);
}

/* Red — martyrs' optional memorials. Cream bg, red accent. */
.palette-red {
  --bg: #FAF7F2; --panel: #F2EDE3; --accent: #9B3535; --gold: #9B3535;
  --text: #1C1714; --dim: #8A7D6E; --shell: #F0EAE0;
  --accent-soft: rgba(155,53,53,0.15);
}

/* White solemnity — pure white background, antique gold accents */
.palette-white {
  --bg: #FEFEFE; --panel: #FEFEFE; --accent: #9AAAB8; --gold: #9AAAB8;
  --text: #1C1714; --dim: #8A8E94; --shell: #EAEBEC;
  --accent-soft: rgba(154,170,184,0.22);
}

/* White Major — TBD (Christmas, Easter). Provisional values matching palette-white. */
.palette-white-major {
  --bg: #FEFEFE; --panel: #FEFEFE; --accent: #9AAAB8; --gold: #9AAAB8;
  --text: #1C1714; --dim: #8A8E94; --shell: #EAEBEC;
  --accent-soft: rgba(154,170,184,0.22);
}

/* Advent — TBD. Provisional purple values. */
.palette-advent {
  --bg: #FAF7F2; --panel: #F2EDE3; --accent: #6B4F9E; --gold: #6B4F9E;
  --text: #1C1714; --dim: #8A7D6E; --shell: #F0EAE0;
  --accent-soft: rgba(107,79,158,0.15);
}

/* Rose — TBD (Gaudete, Laetare). Provisional values. */
.palette-rose {
  --bg: #FAF7F2; --panel: #F2EDE3; --accent: #C4607A; --gold: #C4607A;
  --text: #1C1714; --dim: #8A7D6E; --shell: #F0EAE0;
  --accent-soft: rgba(196,96,122,0.15);
}

/* Black — All Souls. Provisional values. */
.palette-black {
  --bg: #1A1A1A; --panel: #222222; --accent: #888888; --gold: #888888;
  --text: #F0F0F0; --dim: #AAAAAA; --shell: #101010;
  --accent-soft: rgba(136,136,136,0.2);
}

/* ── App shell ── */
.app {
  width: 100%; height: 100svh;
  display: flex; flex-direction: column;
  overflow: hidden; position: relative;
}

/* ── Top bar ── */
.top-bar {
  position: fixed; top: 0; left: 0; right: 0; z-index: 40;
  background: linear-gradient(to bottom, rgba(250,247,242,0.96) 0%, rgba(250,247,242,0) 100%);
  pointer-events: none;
}
.palette-lent .top-bar        { background: linear-gradient(to bottom, rgba(36,30,26,0.96) 0%, rgba(36,30,26,0) 100%); }
.palette-goodfriday .top-bar  { background: linear-gradient(to bottom, rgba(14,12,11,0.96) 0%, rgba(14,12,11,0) 100%); }
.palette-pentecost .top-bar   { background: linear-gradient(to bottom, rgba(139,26,26,0.96) 0%, rgba(139,26,26,0) 100%); }
.palette-ordinarytime .top-bar         { background: linear-gradient(to bottom, rgba(250,247,242,0.96) 0%, rgba(250,247,242,0) 100%); }
.palette-ordinarytime-major .top-bar   { background: linear-gradient(to bottom, rgba(26,43,26,0.96) 0%, rgba(26,43,26,0) 100%); }
.palette-red .top-bar                  { background: linear-gradient(to bottom, rgba(250,247,242,0.96) 0%, rgba(250,247,242,0) 100%); }
.palette-white .top-bar                { background: linear-gradient(to bottom, rgba(254,254,254,0.96) 0%, rgba(254,254,254,0) 100%); }
.palette-white-major .top-bar          { background: linear-gradient(to bottom, rgba(254,254,254,0.96) 0%, rgba(254,254,254,0) 100%); }
.palette-advent .top-bar               { background: linear-gradient(to bottom, rgba(250,247,242,0.96) 0%, rgba(250,247,242,0) 100%); }
.palette-rose .top-bar                 { background: linear-gradient(to bottom, rgba(250,247,242,0.96) 0%, rgba(250,247,242,0) 100%); }
.palette-black .top-bar                { background: linear-gradient(to bottom, rgba(26,26,26,0.96) 0%, rgba(26,26,26,0) 100%); }

.top-bar-inner {
  display: flex;
  align-items: center;
  justify-content: space-between;
  padding: 1.2rem 1.0rem 0.8rem;
  pointer-events: all;
}

.brand-wrap { text-align: center; flex: 1; }
.brand {
  font-family: 'Cormorant Garamond', serif;
  font-size: 1.75rem; font-weight: 400;
  letter-spacing: 0.42em; text-transform: uppercase;
  color: var(--gold); line-height: 1;
}
.day-label {
  font-family: 'Raleway', sans-serif;
  font-size: 0.78rem; font-weight: 500;
  letter-spacing: 0.16em; text-transform: uppercase;
  color: var(--dim); margin-top: 0.25rem;
  line-height: 1;
}

/* ── Screens container ── */
.screens-container {
  position: absolute; top: 0; left: 0;
  width: 100%; height: 100%;
}

.day-screens {
  position: absolute; top: 0; left: 0;
  width: 100%; height: 100%;
}

.screen {
  position: absolute; inset: 0;
  display: flex; flex-direction: column;
  justify-content: flex-start; align-items: flex-start;
  overflow: hidden;
  background: var(--bg);
  opacity: 0;
  pointer-events: none;
  transition: opacity 0.85s ease;
}
.screen.active {
  opacity: 1;
  pointer-events: auto;
}
.screen.leaving {
  opacity: 0;
  pointer-events: none;
  transition: opacity 0.85s ease;
}
.screen-word, .screen-contemplation, .screen-practice { background: var(--panel); }

/* Noise texture */
.screen::before {
  content: '';
  position: absolute; inset: 0;
  background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='300' height='300'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='300' height='300' filter='url(%23n)' opacity='0.025'/%3E%3C/svg%3E");
  pointer-events: none; z-index: 0;
}
.screen > * { position: relative; z-index: 1; }

/* Screen label */
.screen-label {
  font-family: 'Raleway', sans-serif;
  font-size: 1.05rem; font-weight: 600;
  letter-spacing: 0.38em; text-transform: uppercase;
  color: var(--dim);
  padding: 7.2rem 1.4rem 1.2rem;
  width: 100%; flex-shrink: 0;
}

/* Scrollable inner */
.screen-inner {
  flex: 1; overflow-y: auto; overflow-x: hidden;
  -webkit-overflow-scrolling: touch;
  padding: 1.4rem 1.4rem 14rem;
  width: 100%;
  scrollbar-width: thin;
  scrollbar-color: var(--accent) transparent;
}
.screen-inner::-webkit-scrollbar { width: 2px; }
.screen-inner::-webkit-scrollbar-thumb { background: var(--accent); }

/* Fade overlays */
.screen-fade-top, .screen-fade-bottom {
  position: absolute; left: 0; right: 0;
  pointer-events: none; z-index: 2;
}
.screen-fade-top    { top: 0;    height: 14%; }
.screen-fade-bottom { bottom: 0; height: 20%; }

/* Eastertide fades */
.palette-eastertide .screen-season .screen-fade-top,
.palette-eastertide .screen-scripture .screen-fade-top,
.palette-eastertide .screen-prayer .screen-fade-top
  { background: linear-gradient(to bottom, #FAF7F2 40%, rgba(250,247,242,0)); }
.palette-eastertide .screen-season .screen-fade-bottom,
.palette-eastertide .screen-scripture .screen-fade-bottom,
.palette-eastertide .screen-prayer .screen-fade-bottom
  { background: linear-gradient(to top, #FAF7F2 50%, rgba(250,247,242,0)); }
.palette-eastertide .screen-word .screen-fade-top,
.palette-eastertide .screen-contemplation .screen-fade-top,
.palette-eastertide .screen-practice .screen-fade-top
  { background: linear-gradient(to bottom, #F2EDE3 40%, rgba(242,237,227,0)); }
.palette-eastertide .screen-word .screen-fade-bottom,
.palette-eastertide .screen-contemplation .screen-fade-bottom,
.palette-eastertide .screen-practice .screen-fade-bottom
  { background: linear-gradient(to top, #F2EDE3 50%, rgba(242,237,227,0)); }

/* Lent fades */
.palette-lent .screen-season .screen-fade-top,
.palette-lent .screen-scripture .screen-fade-top,
.palette-lent .screen-prayer .screen-fade-top
  { background: linear-gradient(to bottom, #241E1A 40%, rgba(36,30,26,0)); }
.palette-lent .screen-season .screen-fade-bottom,
.palette-lent .screen-scripture .screen-fade-bottom,
.palette-lent .screen-prayer .screen-fade-bottom
  { background: linear-gradient(to top, #241E1A 50%, rgba(36,30,26,0)); }
.palette-lent .screen-word .screen-fade-top,
.palette-lent .screen-contemplation .screen-fade-top,
.palette-lent .screen-practice .screen-fade-top
  { background: linear-gradient(to bottom, #2E2520 40%, rgba(46,37,32,0)); }
.palette-lent .screen-word .screen-fade-bottom,
.palette-lent .screen-contemplation .screen-fade-bottom,
.palette-lent .screen-practice .screen-fade-bottom
  { background: linear-gradient(to top, #2E2520 50%, rgba(46,37,32,0)); }

/* Good Friday fades */
.palette-goodfriday .screen-season .screen-fade-top,
.palette-goodfriday .screen-scripture .screen-fade-top,
.palette-goodfriday .screen-prayer .screen-fade-top
  { background: linear-gradient(to bottom, #0E0C0B 40%, rgba(14,12,11,0)); }
.palette-goodfriday .screen-season .screen-fade-bottom,
.palette-goodfriday .screen-scripture .screen-fade-bottom,
.palette-goodfriday .screen-prayer .screen-fade-bottom
  { background: linear-gradient(to top, #0E0C0B 50%, rgba(14,12,11,0)); }
.palette-goodfriday .screen-word .screen-fade-top,
.palette-goodfriday .screen-contemplation .screen-fade-top,
.palette-goodfriday .screen-practice .screen-fade-top
  { background: linear-gradient(to bottom, #181412 40%, rgba(24,20,18,0)); }
.palette-goodfriday .screen-word .screen-fade-bottom,
.palette-goodfriday .screen-contemplation .screen-fade-bottom,
.palette-goodfriday .screen-practice .screen-fade-bottom
  { background: linear-gradient(to top, #181412 50%, rgba(24,20,18,0)); }

/* Pentecost fades */
.palette-pentecost .screen-season .screen-fade-top,
.palette-pentecost .screen-scripture .screen-fade-top,
.palette-pentecost .screen-prayer .screen-fade-top
  { background: linear-gradient(to bottom, #8B1A1A 40%, rgba(139,26,26,0)); }
.palette-pentecost .screen-season .screen-fade-bottom,
.palette-pentecost .screen-scripture .screen-fade-bottom,
.palette-pentecost .screen-prayer .screen-fade-bottom
  { background: linear-gradient(to top, #8B1A1A 50%, rgba(139,26,26,0)); }
.palette-pentecost .screen-word .screen-fade-top,
.palette-pentecost .screen-contemplation .screen-fade-top,
.palette-pentecost .screen-practice .screen-fade-top
  { background: linear-gradient(to bottom, #7A1515 40%, rgba(122,21,21,0)); }
.palette-pentecost .screen-word .screen-fade-bottom,
.palette-pentecost .screen-contemplation .screen-fade-bottom,
.palette-pentecost .screen-practice .screen-fade-bottom
  { background: linear-gradient(to top, #7A1515 50%, rgba(122,21,21,0)); }

.palette-pentecost .screen-season .screen-fade-top,
.palette-pentecost .screen-scripture .screen-fade-top,
.palette-pentecost .screen-prayer .screen-fade-top
  { background: linear-gradient(to bottom, #8B1A1A 40%, rgba(139,26,26,0)); }
.palette-pentecost .screen-season .screen-fade-bottom,
.palette-pentecost .screen-scripture .screen-fade-bottom,
.palette-pentecost .screen-prayer .screen-fade-bottom
  { background: linear-gradient(to top, #8B1A1A 50%, rgba(139,26,26,0)); }
.palette-pentecost .screen-word .screen-fade-top,
.palette-pentecost .screen-contemplation .screen-fade-top,
.palette-pentecost .screen-practice .screen-fade-top
  { background: linear-gradient(to bottom, #7A1515 40%, rgba(122,21,21,0)); }
.palette-pentecost .screen-word .screen-fade-bottom,
.palette-pentecost .screen-contemplation .screen-fade-bottom,
.palette-pentecost .screen-practice .screen-fade-bottom
  { background: linear-gradient(to top, #7A1515 50%, rgba(122,21,21,0)); }

/* Ordinary Time fades */
.palette-ordinarytime .screen-season .screen-fade-top,
.palette-ordinarytime .screen-scripture .screen-fade-top,
.palette-ordinarytime .screen-prayer .screen-fade-top,
.palette-ordinarytime .screen-amen .screen-fade-top,
.palette-red .screen-season .screen-fade-top,
.palette-red .screen-scripture .screen-fade-top,
.palette-red .screen-prayer .screen-fade-top,
.palette-red .screen-amen .screen-fade-top
  { background: linear-gradient(to bottom, #FAF7F2 40%, rgba(250,247,242,0)); }
.palette-ordinarytime .screen-season .screen-fade-bottom,
.palette-ordinarytime .screen-scripture .screen-fade-bottom,
.palette-ordinarytime .screen-prayer .screen-fade-bottom,
.palette-ordinarytime .screen-amen .screen-fade-bottom,
.palette-red .screen-season .screen-fade-bottom,
.palette-red .screen-scripture .screen-fade-bottom,
.palette-red .screen-prayer .screen-fade-bottom,
.palette-red .screen-amen .screen-fade-bottom
  { background: linear-gradient(to top, #FAF7F2 50%, rgba(250,247,242,0)); }
.palette-ordinarytime .screen-word .screen-fade-top,
.palette-ordinarytime .screen-contemplation .screen-fade-top,
.palette-ordinarytime .screen-practice .screen-fade-top,
.palette-red .screen-word .screen-fade-top,
.palette-red .screen-contemplation .screen-fade-top,
.palette-red .screen-practice .screen-fade-top
  { background: linear-gradient(to bottom, #F2EDE3 40%, rgba(242,237,227,0)); }
.palette-ordinarytime .screen-word .screen-fade-bottom,
.palette-ordinarytime .screen-contemplation .screen-fade-bottom,
.palette-ordinarytime .screen-practice .screen-fade-bottom,
.palette-red .screen-word .screen-fade-bottom,
.palette-red .screen-contemplation .screen-fade-bottom,
.palette-red .screen-practice .screen-fade-bottom
  { background: linear-gradient(to top, #F2EDE3 50%, rgba(242,237,227,0)); }

/* Ordinary Time Major fades */
.palette-ordinarytime-major .screen-season .screen-fade-top,
.palette-ordinarytime-major .screen-scripture .screen-fade-top,
.palette-ordinarytime-major .screen-prayer .screen-fade-top,
.palette-ordinarytime-major .screen-amen .screen-fade-top
  { background: linear-gradient(to bottom, #1A2B1A 40%, rgba(26,43,26,0)); }
.palette-ordinarytime-major .screen-season .screen-fade-bottom,
.palette-ordinarytime-major .screen-scripture .screen-fade-bottom,
.palette-ordinarytime-major .screen-prayer .screen-fade-bottom,
.palette-ordinarytime-major .screen-amen .screen-fade-bottom
  { background: linear-gradient(to top, #1A2B1A 50%, rgba(26,43,26,0)); }
.palette-ordinarytime-major .screen-word .screen-fade-top,
.palette-ordinarytime-major .screen-contemplation .screen-fade-top,
.palette-ordinarytime-major .screen-practice .screen-fade-top
  { background: linear-gradient(to bottom, #162414 40%, rgba(22,36,20,0)); }
.palette-ordinarytime-major .screen-word .screen-fade-bottom,
.palette-ordinarytime-major .screen-contemplation .screen-fade-bottom,
.palette-ordinarytime-major .screen-practice .screen-fade-bottom
  { background: linear-gradient(to top, #162414 50%, rgba(22,36,20,0)); }

/* White fades */
.palette-white .screen-season .screen-fade-top,
.palette-white .screen-scripture .screen-fade-top,
.palette-white .screen-prayer .screen-fade-top,
.palette-white .screen-amen .screen-fade-top,
.palette-white-major .screen-season .screen-fade-top,
.palette-white-major .screen-scripture .screen-fade-top,
.palette-white-major .screen-prayer .screen-fade-top,
.palette-white-major .screen-amen .screen-fade-top
  { background: linear-gradient(to bottom, #FEFEFE 40%, rgba(254,254,254,0)); }
.palette-white .screen-season .screen-fade-bottom,
.palette-white .screen-scripture .screen-fade-bottom,
.palette-white .screen-prayer .screen-fade-bottom,
.palette-white .screen-amen .screen-fade-bottom,
.palette-white-major .screen-season .screen-fade-bottom,
.palette-white-major .screen-scripture .screen-fade-bottom,
.palette-white-major .screen-prayer .screen-fade-bottom,
.palette-white-major .screen-amen .screen-fade-bottom
  { background: linear-gradient(to top, #FEFEFE 50%, rgba(254,254,254,0)); }
.palette-white .screen-word .screen-fade-top,
.palette-white .screen-contemplation .screen-fade-top,
.palette-white .screen-practice .screen-fade-top,
.palette-white-major .screen-word .screen-fade-top,
.palette-white-major .screen-contemplation .screen-fade-top,
.palette-white-major .screen-practice .screen-fade-top
  { background: linear-gradient(to bottom, #FEFEFE 40%, rgba(254,254,254,0)); }
.palette-white .screen-word .screen-fade-bottom,
.palette-white .screen-contemplation .screen-fade-bottom,
.palette-white .screen-practice .screen-fade-bottom,
.palette-white-major .screen-word .screen-fade-bottom,
.palette-white-major .screen-contemplation .screen-fade-bottom,
.palette-white-major .screen-practice .screen-fade-bottom
  { background: linear-gradient(to top, #FEFEFE 50%, rgba(254,254,254,0)); }

/* Advent fades (provisional — cream bg) */
.palette-advent .screen-season .screen-fade-top,
.palette-advent .screen-scripture .screen-fade-top,
.palette-advent .screen-prayer .screen-fade-top,
.palette-advent .screen-amen .screen-fade-top
  { background: linear-gradient(to bottom, #FAF7F2 40%, rgba(250,247,242,0)); }
.palette-advent .screen-season .screen-fade-bottom,
.palette-advent .screen-scripture .screen-fade-bottom,
.palette-advent .screen-prayer .screen-fade-bottom,
.palette-advent .screen-amen .screen-fade-bottom
  { background: linear-gradient(to top, #FAF7F2 50%, rgba(250,247,242,0)); }
.palette-advent .screen-word .screen-fade-top,
.palette-advent .screen-contemplation .screen-fade-top,
.palette-advent .screen-practice .screen-fade-top
  { background: linear-gradient(to bottom, #F2EDE3 40%, rgba(242,237,227,0)); }
.palette-advent .screen-word .screen-fade-bottom,
.palette-advent .screen-contemplation .screen-fade-bottom,
.palette-advent .screen-practice .screen-fade-bottom
  { background: linear-gradient(to top, #F2EDE3 50%, rgba(242,237,227,0)); }

/* Rose fades (provisional — cream bg) */
.palette-rose .screen-season .screen-fade-top,
.palette-rose .screen-scripture .screen-fade-top,
.palette-rose .screen-prayer .screen-fade-top,
.palette-rose .screen-amen .screen-fade-top
  { background: linear-gradient(to bottom, #FAF7F2 40%, rgba(250,247,242,0)); }
.palette-rose .screen-season .screen-fade-bottom,
.palette-rose .screen-scripture .screen-fade-bottom,
.palette-rose .screen-prayer .screen-fade-bottom,
.palette-rose .screen-amen .screen-fade-bottom
  { background: linear-gradient(to top, #FAF7F2 50%, rgba(250,247,242,0)); }
.palette-rose .screen-word .screen-fade-top,
.palette-rose .screen-contemplation .screen-fade-top,
.palette-rose .screen-practice .screen-fade-top
  { background: linear-gradient(to bottom, #F2EDE3 40%, rgba(242,237,227,0)); }
.palette-rose .screen-word .screen-fade-bottom,
.palette-rose .screen-contemplation .screen-fade-bottom,
.palette-rose .screen-practice .screen-fade-bottom
  { background: linear-gradient(to top, #F2EDE3 50%, rgba(242,237,227,0)); }

/* Black fades (All Souls) */
.palette-black .screen-season .screen-fade-top,
.palette-black .screen-scripture .screen-fade-top,
.palette-black .screen-prayer .screen-fade-top,
.palette-black .screen-amen .screen-fade-top
  { background: linear-gradient(to bottom, #1A1A1A 40%, rgba(26,26,26,0)); }
.palette-black .screen-season .screen-fade-bottom,
.palette-black .screen-scripture .screen-fade-bottom,
.palette-black .screen-prayer .screen-fade-bottom,
.palette-black .screen-amen .screen-fade-bottom
  { background: linear-gradient(to top, #1A1A1A 50%, rgba(26,26,26,0)); }
.palette-black .screen-word .screen-fade-top,
.palette-black .screen-contemplation .screen-fade-top,
.palette-black .screen-practice .screen-fade-top
  { background: linear-gradient(to bottom, #222222 40%, rgba(34,34,34,0)); }
.palette-black .screen-word .screen-fade-bottom,
.palette-black .screen-contemplation .screen-fade-bottom,
.palette-black .screen-practice .screen-fade-bottom
  { background: linear-gradient(to top, #222222 50%, rgba(34,34,34,0)); }

/* Desktop nav arrows — hidden by default */
.desktop-nav-arrow { display: none; }

/* ── Threshold veil: Screen 7 → 1 wrap ── */
#thresholdVeil {
  position: fixed; inset: 0; z-index: 60;
  background: #100D0B;
  opacity: 0; pointer-events: none;
  transition: opacity 0.6s ease;
}
#thresholdVeil.threshold-closing { opacity: 1; transition: opacity 0.6s ease; }
#thresholdVeil.threshold-opening { opacity: 0; transition: opacity 0.7s ease; }

@keyframes fadeUp {
  from { opacity: 0; transform: translateY(20px); }
  to   { opacity: 1; transform: translateY(0); }
}
.screen.animating .screen-label                  { animation: fadeUp 0.5s ease forwards; animation-delay: 0.10s; opacity: 0; }
.screen.animating .screen-inner > *:nth-child(1) { animation: fadeUp 0.5s ease forwards; animation-delay: 0.18s; opacity: 0; }
.screen.animating .screen-inner > *:nth-child(2) { animation: fadeUp 0.5s ease forwards; animation-delay: 0.25s; opacity: 0; }
.screen.animating .screen-inner > *:nth-child(3) { animation: fadeUp 0.5s ease forwards; animation-delay: 0.32s; opacity: 0; }
.screen.animating .screen-inner > *:nth-child(4) { animation: fadeUp 0.5s ease forwards; animation-delay: 0.39s; opacity: 0; }
.screen.animating .screen-inner > *:nth-child(5) { animation: fadeUp 0.5s ease forwards; animation-delay: 0.46s; opacity: 0; }

/* ── Screen 1: Season ── */
.season-day {
  font-family: 'Cormorant Garamond', serif;
  font-size: clamp(1.8rem, 9vw, 3rem); font-weight: 300;
  letter-spacing: 0.18em; text-transform: uppercase;
  line-height: 1.2; color: var(--text);
  margin-bottom: 0.5rem; white-space: normal; word-break: break-word; max-width: 90vw;
}
.season-week {
  font-family: 'Raleway', sans-serif;
  font-size: 0.95rem; font-weight: 400;
  letter-spacing: 0.15em; color: var(--dim);
  margin-bottom: 0.4rem; line-height: 1.6;
}
.season-feast {
  font-family: 'Raleway', sans-serif;
  font-size: 0.95rem; font-weight: 400;
  letter-spacing: 0.14em; color: var(--gold); line-height: 1.6;
}
.season-divider {
  width: 100%; max-width: 420px; height: 1px;
  background: linear-gradient(to right, var(--gold), transparent);
  margin: 1.8rem 0; flex-shrink: 0;
}
.season-body {
  font-family: 'Libre Baskerville', serif;
  font-size: 1.05rem; line-height: 2;
  max-width: 420px; color: var(--text);
}
.season-body p + p { margin-top: 1.1rem; }

/* ── Screen 2: Word ── */
.word-title {
  font-family: 'Cormorant Garamond', serif;
  font-size: 2rem; font-weight: 300;
  letter-spacing: 0.12em; text-transform: uppercase;
  line-height: 1.1; color: var(--text);
  margin-bottom: 0.75rem;
  white-space: nowrap;
  display: block; max-width: 100%;
}
.word-pronunciation {
  font-family: 'Raleway', sans-serif;
  font-size: 0.95rem; font-weight: 700; font-style: normal;
  letter-spacing: 0.18em; color: var(--dim);
  margin-bottom: 1.8rem; line-height: 1.6;
}
.word-pronunciation + .word-divider { margin-top: 0; }
.word-divider {
  width: 100%; max-width: 420px; height: 1px;
  background: linear-gradient(to right, var(--gold), transparent);
  margin: 0 0 1.8rem 0; flex-shrink: 0;
}
.word-definition {
  font-family: 'Libre Baskerville', serif;
  font-size: 1.05rem; line-height: 2;
  max-width: 420px; color: var(--text);
}
.word-definition p + p { margin-top: 1.1rem; }
.word-definition em { font-style: italic; }

/* ── Screen 3: Scripture ── */
.scripture-inter-verse { height: 1.8rem; }
.scripture-block {
  position: relative; padding-left: 1.5rem; max-width: 420px;
}
.scripture-block::before {
  content: '';
  position: absolute; left: 0; top: 0; bottom: 0;
  width: 2px;
  background: linear-gradient(to bottom, var(--gold), transparent);
  pointer-events: none;
}
.scripture-text {
  font-family: 'Cormorant Garamond', serif;
  font-size: 1.35rem; font-style: italic; font-weight: 500;
  line-height: 2.1; color: var(--text);
}
.scripture-text p + p { margin-top: 1rem; }
.scripture-translation {
  font-family: 'Raleway', sans-serif;
  font-size: 0.85rem; font-weight: 400;
  letter-spacing: 0.18em; text-transform: uppercase;
  color: var(--dim); margin-top: 1.8rem;
}

.screen-scripture .screen-inner > *:first-child { margin-top: 1.2rem; }
.screen-contemplation .screen-inner > *:first-child { margin-top: 1.2rem; }
.screen-prayer .screen-inner > *:first-child { margin-top: 1.2rem; }
.screen-practice .screen-inner > *:first-child { margin-top: 1.2rem; }

/* ── Screen 4: Contemplation ── */
.contemplation-body {
  font-family: 'Libre Baskerville', serif;
  font-size: 1.05rem; line-height: 2;
  max-width: 420px; color: var(--text);
}
.contemplation-body p + p { margin-top: 1.1rem; }
.contemplation-question {
  font-family: 'Cormorant Garamond', serif;
  font-size: 1.3rem; font-style: italic; font-weight: 500;
  line-height: 1.65; max-width: 420px;
  margin-top: 2rem; padding-top: 1.5rem;
  border-top: 1px solid transparent;
  border-image: linear-gradient(to right, var(--gold), transparent) 1;
  color: var(--text);
}

/* ── Screen 5: Prayer ── */
.prayer-block {
  position: relative; padding-left: 1.5rem; max-width: 420px;
}
.prayer-block::before {
  content: '';
  position: absolute; left: 0; top: 0; bottom: 0;
  width: 2px;
  background: linear-gradient(to bottom, var(--gold), transparent);
  pointer-events: none;
}
.prayer-body {
  font-family: 'Cormorant Garamond', serif;
  font-size: 1.3rem; font-style: italic; font-weight: 500;
  line-height: 1.75; color: var(--text);
  margin-top: 1.2rem;
}
.prayer-body p + p { margin-top: 0.8rem; }
.prayer-amen {
  font-family: 'Cormorant Garamond', serif;
  font-size: 1.3rem; font-style: italic; font-weight: 500;
  letter-spacing: 0.05em; color: var(--text); margin-top: 1.2rem;
}

/* ── Screen 6: Practice ── */
.practice-body {
  font-family: 'Libre Baskerville', serif;
  font-size: 1.05rem; line-height: 2;
  max-width: 420px; color: var(--text);
}
.practice-body p + p { margin-top: 1.1rem; }
.practice-anchor {
  font-family: 'Cormorant Garamond', serif;
  font-size: 1.3rem; font-style: italic; font-weight: 500;
  line-height: 1.65; max-width: 420px;
  margin-top: 2rem; padding-top: 1.5rem;
  border-top: 1px solid transparent;
  border-image: linear-gradient(to right, var(--gold), transparent) 1;
  color: var(--text);
}

/* ── Figure links ── */
.figure-link {
  background: none; border: none; padding: 0;
  font: inherit; color: inherit; cursor: pointer;
  text-decoration: underline;
  text-decoration-color: var(--gold);
  text-underline-offset: 3px;
  text-decoration-thickness: 1.5px;
  font-weight: inherit;
  transition: color 0.2s ease, text-decoration-color 0.2s ease;
}
.figure-link:hover, .figure-link:focus-visible {
  color: var(--accent); text-decoration-color: var(--accent); outline: none;
}

/* ── Navigation dots ── */
.nav-dots {
  position: fixed; bottom: 1.8rem; left: 50%; transform: translateX(-50%);
  display: flex; flex-direction: row; gap: 0.7rem; z-index: 30;
}
.dot {
  width: 8px; height: 8px; border-radius: 50%;
  background: var(--dim); cursor: pointer;
  transition: all 0.3s ease; border: none; padding: 0; opacity: 0.6;
}
.dot.active { background: var(--gold); transform: scale(1.5); opacity: 1; }

.screen-counter {
  position: fixed; bottom: 1.9rem; right: 1.5rem;
  font-family: 'Raleway', sans-serif; font-size: 0.65rem;
  font-weight: 400; letter-spacing: 0.2em; color: var(--dim); z-index: 30;
}
.swipe-hint {
  position: fixed; bottom: 1.85rem; left: 1.5rem;
  font-family: 'Raleway', sans-serif; font-size: 0.7rem;
  font-weight: 500; letter-spacing: 0.15em; text-transform: uppercase;
  color: var(--dim); opacity: 0.7; z-index: 30;
  transition: opacity 0.5s ease;
}

/* ── Install nudge ── */
.install-nudge {
  position: fixed; bottom: 4.8rem; left: 0; right: 0;
  padding: 0 1.4rem;
  z-index: 30;
  pointer-events: none;
  transition: filter 0.3s ease, opacity 0.3s ease;
}
.install-nudge-rule {
  border: none;
  border-top: 1px solid var(--accent-soft);
  margin-bottom: 0.55rem;
}
.install-nudge-text {
  font-family: 'Libre Baskerville', serif;
  font-size: 0.88rem;
  font-style: italic;
  color: var(--dim);
  pointer-events: none;
}
.install-nudge-link {
  color: var(--gold);
  text-decoration: underline;
  text-underline-offset: 3px;
  text-decoration-thickness: 1px;
  cursor: pointer;
  background: none;
  border: none;
  font-family: 'Libre Baskerville', serif;
  font-size: 0.88rem;
  font-style: italic;
  color: var(--gold);
  padding: 0;
  pointer-events: auto;
  -webkit-tap-highlight-color: transparent;
}
.install-nudge-link:hover { opacity: 0.8; }
.install-modal-open .install-nudge {
  filter: blur(2px);
  opacity: 0.4;
  pointer-events: none;
}

/* ── Install modal ── */
.install-modal-overlay {
  position: fixed; inset: 0;
  background: rgba(200,190,170,0.55);
  display: flex; align-items: flex-end; justify-content: center;
  z-index: 110;
  opacity: 0; pointer-events: none;
}
.install-modal-overlay.install-modal-visible {
  opacity: 1; pointer-events: auto;
}
.palette-lent .install-modal-overlay,
.palette-goodfriday .install-modal-overlay,
.palette-pentecost .install-modal-overlay { background: rgba(5,3,8,0.72); }
.install-modal-card {
  background: var(--bg);
  border: 1px solid var(--accent-soft);
  border-bottom: none;
  width: 100%;
  max-width: 480px;
  padding: 2rem 1.8rem 2.4rem;
  position: relative;
  box-shadow: 0 -12px 48px rgba(28,23,20,0.14);
  transform: translateY(100%);
}
.install-modal-close {
  position: absolute; top: 1rem; right: 1.2rem;
  background: none; border: none;
  font-family: 'Cormorant Garamond', serif;
  font-size: 2rem; font-weight: 300;
  color: var(--dim); line-height: 1;
  cursor: pointer; padding: 0.2rem 0.4rem;
  transition: color 0.2s ease;
  -webkit-tap-highlight-color: transparent;
}
.install-modal-close:hover { color: var(--text); }
.install-modal-wordmark {
  font-family: 'Cormorant Garamond', serif;
  font-size: 1.75rem; font-weight: 400;
  letter-spacing: 0.42em; text-transform: uppercase;
  color: var(--gold); margin-bottom: 1.1rem;
}
.install-modal-intro {
  font-family: 'Libre Baskerville', serif;
  font-size: 1.05rem; line-height: 2;
  color: var(--text); margin-bottom: 1.3rem;
}
.install-modal-steps {
  display: flex; flex-direction: column; gap: 0.9rem;
}
.install-modal-step {
  display: flex; align-items: flex-start; gap: 1rem;
}
.install-modal-step-num {
  font-family: 'Raleway', sans-serif;
  font-size: 0.78rem; font-weight: 600;
  letter-spacing: 0.1em; color: var(--gold);
  margin-top: 0.2rem; flex-shrink: 0;
  width: 1rem; text-align: right;
}
.install-modal-step-text {
  font-family: 'Libre Baskerville', serif;
  font-size: 1.05rem; line-height: 2;
  color: var(--text);
}
.install-modal-step-text strong { font-weight: 700; }
.install-dots-pill {
  display: inline-block;
  font-family: -apple-system, 'SF Pro Text', sans-serif;
  font-size: 0.95rem; font-weight: 500;
  background: rgba(184,134,11,0.1); color: var(--gold);
  padding: 0.1rem 0.5rem; border-radius: 4px;
  letter-spacing: 0.05em; line-height: 1.5;
  vertical-align: baseline;
}
.install-modal-rule {
  border: none;
  border-top: 1px solid var(--accent-soft);
  margin: 1.2rem 0 1.1rem;
}
.install-modal-footnote {
  font-family: 'Libre Baskerville', serif;
  font-size: 1.05rem; font-style: italic;
  color: var(--dim); line-height: 2;
}

/* ── Figure modal ── */
.figure-overlay {
  position: fixed; inset: 0;
  background: rgba(200,190,170,0.6);
  backdrop-filter: blur(6px); -webkit-backdrop-filter: blur(6px);
  z-index: 100; display: flex;
  align-items: center; justify-content: center; padding: 1.5rem;
  opacity: 0; pointer-events: none;
  transition: opacity 0.75s ease;
}
.palette-lent .figure-overlay,
.palette-goodfriday .figure-overlay,
.palette-pentecost .figure-overlay { background: rgba(60,8,8,0.82); }
.palette-ordinarytime-major .figure-overlay { background: rgba(5,12,5,0.82); }
.palette-black .figure-overlay              { background: rgba(0,0,0,0.88); }
.figure-overlay.open { opacity: 1; pointer-events: auto; }

/* ── First-visit welcome overlay ── */
.welcome-overlay {
  position: fixed; inset: 0;
  background: rgba(200,190,170,0.6);
  backdrop-filter: blur(6px); -webkit-backdrop-filter: blur(6px);
  z-index: 100;
  display: flex; align-items: center; justify-content: center; padding: 1.5rem;
  opacity: 0; animation: welcomeFadeIn 1s ease 1.5s forwards;
}
.palette-lent .welcome-overlay,
.palette-goodfriday .welcome-overlay,
.palette-pentecost .welcome-overlay { background: rgba(60,8,8,0.82); }
.palette-ordinarytime-major .welcome-overlay { background: rgba(5,12,5,0.82); }
.palette-black .welcome-overlay              { background: rgba(0,0,0,0.88); }
.welcome-overlay.welcome-hiding {
  animation: welcomeFadeOut 0.3s ease forwards;
}
@keyframes welcomeFadeIn  { from { opacity: 0; } to { opacity: 1; } }
@keyframes welcomeFadeOut { from { opacity: 1; } to { opacity: 0; } }
.welcome-card {
  background: var(--bg); border: 1px solid var(--accent-soft);
  max-width: 400px; width: 100%;
  padding: 2.8rem 2rem 2.4rem;
  text-align: center;
  box-shadow: 0 16px 48px rgba(28,23,20,0.18);
}
.welcome-wordmark {
  font-family: 'Cormorant Garamond', serif;
  font-size: 1.75rem; font-weight: 300;
  letter-spacing: 0.38em; text-transform: uppercase;
  color: var(--gold); margin-bottom: 1.8rem;
}
.welcome-body {
  font-family: 'Libre Baskerville', serif;
  font-size: 1.05rem; font-weight: 400; line-height: 2;
  color: var(--text); margin-bottom: 2rem;
}
.welcome-begin {
  display: inline-block;
  font-family: 'Raleway', sans-serif; font-size: 0.62rem;
  font-weight: 600; letter-spacing: 0.28em; text-transform: uppercase;
  color: var(--bg); background: var(--gold);
  border: none; padding: 1.1rem 2.4rem;
  cursor: pointer; transition: opacity 0.2s;
  -webkit-tap-highlight-color: transparent;
}
.welcome-begin:hover { opacity: 0.85; }
.figure-card {
  background: var(--bg); border: 1px solid var(--accent-soft);
  max-width: 400px; width: 100%; max-height: 80vh; overflow-y: auto;
  padding: 2rem 1.8rem 1.8rem; position: relative;
  box-shadow: 0 16px 48px rgba(28,23,20,0.18);
  text-transform: none; letter-spacing: normal;
}
.figure-card .close {
  position: absolute; top: 0.8rem; right: 1rem;
  background: none; border: none; color: var(--dim);
  font-family: 'Cormorant Garamond', serif;
  font-size: 1.5rem; cursor: pointer; line-height: 1; padding: 0.25rem 0.5rem;
}
.figure-card .close:hover { color: var(--text); }
.figure-eyebrow {
  font-family: 'Raleway', sans-serif; font-size: 0.55rem;
  font-weight: 600; letter-spacing: 0.38em; text-transform: uppercase;
  color: var(--gold); margin-bottom: 0.9rem;
}
.figure-name {
  font-family: 'Cormorant Garamond', serif; font-size: 1.7rem;
  font-weight: 400; letter-spacing: 0.04em;
  color: var(--text); margin-bottom: 0.3rem; line-height: 1.2;
}
.figure-dates {
  font-family: 'Libre Baskerville', serif; font-size: 0.9rem;
  font-style: italic; color: var(--dim); margin-bottom: 0.8rem;
  padding-bottom: 0.8rem;
  border-bottom: 1px solid transparent;
  border-image: linear-gradient(to right, var(--gold), transparent) 1;
}
.figure-role {
  font-family: 'Raleway', sans-serif; font-size: 0.65rem;
  font-weight: 400; letter-spacing: 0.14em; text-transform: uppercase;
  color: var(--dim); margin-bottom: 1.1rem; opacity: 0.8;
}
.figure-divider { display: none; }
.figure-body {
  font-family: 'Libre Baskerville', serif; font-size: 0.95rem;
  line-height: 1.9; color: var(--text);
  text-transform: none; letter-spacing: normal;
}

@media (min-width: 700px) {
  .screen-label { padding: 7.2rem 3rem 1.2rem; }
  .screen-inner { padding-left: 3rem; padding-right: 3rem; }
}

/* ── Scroll overflow indicator ── */
.scroll-indicator {
  display: flex;
  position: absolute;
  bottom: 4.5rem;
  left: 50%;
  transform: translateX(-50%);
  z-index: 5;
  pointer-events: none;
  opacity: 0;
  transition: opacity 0.4s ease;
  animation: scrollBounce 1.6s ease-in-out infinite;
}
.scroll-indicator.visible { opacity: 1; }
.scroll-indicator svg { color: var(--gold); }
@keyframes scrollBounce {
  0%, 100% { transform: translateX(-50%) translateY(0); }
  50%       { transform: translateX(-50%) translateY(5px); }
}

@media (min-width: 680px) {
  html { overflow: auto; overflow-x: hidden; }
  body { overflow: auto; overflow-x: hidden; }

  .app {
    max-width: 480px;
    margin: 0 auto;
    height: 100svh;
    position: relative;
    box-shadow: 0 0 0 100vmax var(--shell);
  }

  .screen { min-width: 100%; width: 100%; }

  .top-bar {
    max-width: 480px;
    left: 50%; transform: translateX(-50%);
  }

  .nav-dots { gap: 0.9rem; }
  .dot { width: 10px; height: 10px; }
  .dot.active { transform: scale(1.5); }

  .swipe-hint { display: none; }

  .desktop-nav-arrow {
    display: flex;
    position: fixed;
    top: 50%;
    transform: translateY(-50%);
    z-index: 35;
    background: none;
    border: none;
    cursor: pointer;
    align-items: center;
    justify-content: center;
    color: var(--gold);
    font-family: 'Cormorant Garamond', serif;
    font-size: 3.5rem;
    font-weight: 300;
    line-height: 1;
    padding: 0.5rem 1rem;
    opacity: 0.55;
    transition: opacity 0.2s ease, color 0.2s ease;
    -webkit-tap-highlight-color: transparent;
  }
  .desktop-nav-arrow:hover { opacity: 1; color: var(--accent); }
  .desktop-nav-arrow.disabled { opacity: 0.15; cursor: default; pointer-events: none; }
  .desktop-nav-prev { left: calc(50% - 240px - 4rem); }
  .desktop-nav-next { right: calc(50% - 240px - 4rem); }

  .screen-amen .screen-inner {
    overflow-y: auto;
    -webkit-overflow-scrolling: touch;
  }

  .screen-fade-bottom { height: 28%; }
}

/* ══════════════════════════════════════════════════════════
   Screen 7 — Amen
   ══════════════════════════════════════════════════════════ */

.screen-amen { background: var(--bg); }

.screen-amen .screen-inner {
  padding: 1.4rem 1.4rem 0;
  display: block;
  overflow-y: auto;
  -webkit-overflow-scrolling: touch;
}
.screen-amen .screen-inner > * {
  margin-left: auto;
  margin-right: auto;
  width: 100%;
  max-width: 100%;
}
.amen-spacer { height: 12rem; flex-shrink: 0; }

.amen-hero {
  display: flex; flex-direction: column; align-items: center;
  text-align: center; width: 100%;
  padding: 0.5rem 0 1.8rem;
  margin-bottom: 0;
  border-bottom: 1px solid transparent;
  transition: border-color 0.9s ease;
}
.amen-hero.amen-hero-checked {
  border-bottom-color: var(--accent-soft);
}

.amen-checkbox-row {
  display: flex; align-items: center; justify-content: center;
  gap: 1rem; cursor: pointer; user-select: none;
  -webkit-tap-highlight-color: transparent;
  margin-bottom: 1.8rem; width: 100%;
}
.amen-checkbox-outer {
  width: 32px; height: 32px; flex-shrink: 0;
  border: 1.5px solid var(--accent-soft);
  display: flex; align-items: center; justify-content: center;
  transition: border-color 0.25s ease, background 0.25s ease;
  color: transparent;
}
.amen-checkbox-outer.amen-checked {
  border-color: var(--gold);
  background: var(--accent-soft);
  color: var(--gold);
}
.amen-tick {
  opacity: 0; transform: scale(0.6);
  transition: opacity 0.2s ease, transform 0.25s cubic-bezier(0.34,1.56,0.64,1);
}
.amen-checked .amen-tick { opacity: 1; transform: scale(1); }
.amen-checkbox-label {
  font-family: 'Cormorant Garamond', serif;
  font-size: clamp(1.5rem, 6vw, 1.9rem);
  font-weight: 300; font-style: italic;
  letter-spacing: 0.02em; color: var(--text); line-height: 1.15;
  transition: color 0.2s ease;
}
.amen-checkbox-row:hover .amen-checkbox-label { color: var(--accent); }

.amen-word-block {
  display: flex; flex-direction: column; align-items: center; gap: 0.35rem;
  width: 100%;
  opacity: 0; transform: translateY(-4px);
  max-height: 0; overflow: hidden; margin-bottom: 0;
  transition: opacity 0.9s ease, transform 0.9s ease,
              max-height 0.9s ease, margin-bottom 0.9s ease;
}
.amen-word-block.amen-visible {
  opacity: 1; transform: translateY(0);
  max-height: 300px; margin-bottom: 0;
}
.amen-word-rule {
  width: 2rem; height: 1px;
  background: linear-gradient(to right, transparent, var(--gold), transparent);
  margin-bottom: 0.5rem;
}
.amen-word-preposition {
  font-family: 'Raleway', sans-serif; font-size: 0.85rem; font-weight: 400;
  letter-spacing: 0.18em; text-transform: uppercase; color: var(--dim);
}
.amen-word-name {
  font-family: 'Cormorant Garamond', serif;
  font-size: clamp(2.6rem, 13vw, 3.6rem);
  font-weight: 600; letter-spacing: 0.14em; text-transform: uppercase;
  color: var(--text); line-height: 1;
  white-space: nowrap; display: block; max-width: 100%;
}
.amen-word-date {
  font-family: 'Raleway', sans-serif; font-size: 0.9rem; font-weight: 400;
  letter-spacing: 0.12em; color: var(--dim); margin-top: 0.1rem;
}

.amen-commission {
  width: 100%; text-align: center;
  padding: 0; max-height: 0; overflow: hidden;
  border-top: 1px solid transparent;
  border-bottom: 1px solid transparent;
  margin-bottom: 0;
  opacity: 0; transform: translateY(-4px);
  transition: opacity 0.9s ease, transform 0.9s ease,
              max-height 0.9s ease, padding 0.9s ease, margin-bottom 0.9s ease,
              border-top-color 0.9s ease, border-bottom-color 0.9s ease;
  pointer-events: none;
}
.amen-commission.amen-visible {
  opacity: 1; transform: translateY(0);
  max-height: 300px; padding: 1.6rem 0 1.4rem; margin-bottom: 1.6rem;
  border-top-color: var(--accent-soft);
  border-bottom-color: var(--accent-soft);
  pointer-events: auto;
}
.amen-commission-line1 {
  font-family: 'Cormorant Garamond', serif;
  font-size: clamp(1.2rem, 5vw, 1.45rem);
  font-weight: 400; font-style: italic;
  color: var(--text); line-height: 1.5;
}
.amen-commission-line1 + .amen-commission-line1 { margin-top: 0.2rem; }

.amen-count-block {
  display: none; text-align: center; width: 100%; margin-bottom: 1.5rem;
}
.amen-count-block.amen-count-visible { display: block; }
.amen-count-number {
  font-family: 'Cormorant Garamond', serif;
  font-size: clamp(2.2rem, 10vw, 2.8rem); font-weight: 300;
  letter-spacing: 0.04em; color: var(--text); line-height: 1; margin-bottom: 0.6rem;
}
.amen-count-label {
  font-family: 'Libre Baskerville', serif; font-size: 1rem; font-style: italic;
  color: var(--dim); line-height: 1.75; max-width: 300px; margin: 0 auto;
}

.amen-share-block {
  width: 100%; text-align: center;
  opacity: 0; transform: translateY(-4px);
  max-height: 0; overflow: hidden; margin-bottom: 0;
  transition: opacity 0.9s ease, transform 0.9s ease,
              max-height 0.9s ease, margin-bottom 0.9s ease;
  pointer-events: none;
}
.amen-share-block.amen-visible {
  opacity: 1; transform: translateY(0);
  max-height: 700px; margin-bottom: 1.6rem;
  pointer-events: all;
}
.amen-share-prompt {
  font-family: 'Libre Baskerville', serif; font-size: 1rem; font-style: italic;
  color: var(--dim); line-height: 1.75; margin-bottom: 1.4rem; text-align: center;
}
.amen-share-btn-primary {
  display: block; width: 100%;
  font-family: 'Raleway', sans-serif; font-size: 0.75rem; font-weight: 600;
  letter-spacing: 0.34em; text-transform: uppercase;
  color: var(--bg); background: var(--gold);
  border: none; padding: 1.1rem 1.5rem;
  cursor: pointer; margin-bottom: 0.8rem;
  transition: opacity 0.2s; -webkit-tap-highlight-color: transparent;
}
.amen-share-btn-primary:hover { opacity: 0.85; }
.amen-share-btn-secondary {
  display: block; width: 100%;
  font-family: 'Raleway', sans-serif; font-size: 0.75rem; font-weight: 600;
  letter-spacing: 0.34em; text-transform: uppercase;
  color: var(--gold); background: none;
  border: 1px solid var(--accent-soft); padding: 0.95rem 1.5rem;
  cursor: pointer; margin-bottom: 0.8rem;
  transition: opacity 0.2s, border-color 0.2s; -webkit-tap-highlight-color: transparent;
}
.amen-share-btn-secondary:hover { opacity: 0.75; border-color: var(--gold); }
.amen-not-now-btn {
  display: block; width: 100%;
  font-family: 'Libre Baskerville', serif; font-size: 0.95rem; font-style: italic;
  color: var(--dim); background: none; border: none;
  padding: 0.5rem 0; text-align: center; cursor: pointer;
  transition: color 0.2s; -webkit-tap-highlight-color: transparent;
}
.amen-not-now-btn:hover { color: var(--text); }
.amen-share-confirmed {
  display: none; font-family: 'Libre Baskerville', serif;
  font-size: 1rem; font-style: italic; color: var(--dim);
  line-height: 1.75; text-align: center;
}
.amen-share-confirmed.amen-share-confirmed-visible { display: block; }

.amen-invitation-section {
  width: 100%; text-align: center;
  padding: 1.4rem 0 0;
  border-top: 1px solid var(--accent-soft);
}
.amen-invitation-trigger {
  display: block; width: 100%;
  cursor: pointer; -webkit-tap-highlight-color: transparent;
  background: none; border: none; padding: 0;
  font: inherit; text-decoration: none;
  -webkit-user-select: none; user-select: none;
  touch-action: manipulation;
}
.amen-invitation-trigger-inner {
  display: flex; align-items: center; justify-content: center; gap: 0.45rem;
  padding: 1.1rem 1rem;
  border-bottom: 1px solid var(--accent-soft);
  transition: border-color 0.2s;
}
.amen-invitation-trigger:hover .amen-invitation-trigger-inner { border-color: var(--gold); }
.amen-invitation-trigger:hover .amen-invitation-text { color: var(--accent); }
.amen-invitation-text {
  font-family: 'Cormorant Garamond', serif; font-size: 1.2rem;
  font-weight: 400; font-style: italic; color: var(--text); line-height: 1.4;
  transition: color 0.2s;
}
.amen-invitation-arrow {
  font-size: 1.4rem; color: var(--gold); line-height: 1; font-style: normal;
  display: inline-block;
  transition: transform 0.3s cubic-bezier(0.4,0,0.2,1);
  transform-origin: 45% 50%;
}
.amen-invitation-arrow.amen-inv-arrow-open { transform: rotate(90deg); }

.amen-invitation-panel {
  overflow: hidden; max-height: 0; opacity: 0;
  transition: max-height 0.45s cubic-bezier(0.4,0,0.2,1), opacity 0.3s ease 0.08s;
  width: 100%;
}
.amen-invitation-panel.amen-inv-open { max-height: 600px; opacity: 1; }
.amen-invitation-panel-inner {
  padding: 1.3rem 0 0.5rem;
  display: flex; flex-direction: column; align-items: center; gap: 1rem;
}
.amen-invitation-message {
  font-family: 'Libre Baskerville', serif; font-size: 0.95rem; font-style: italic;
  color: var(--dim); line-height: 1.85;
  border: 1px solid var(--accent-soft); padding: 1.1rem 1.2rem;
  background: var(--accent-soft); width: 100%; text-align: center;
}
.amen-invitation-send {
  display: block; width: 100%;
  font-family: 'Raleway', sans-serif; font-size: 0.62rem; font-weight: 600;
  letter-spacing: 0.28em; text-transform: uppercase;
  color: var(--bg); background: var(--gold);
  border: none; padding: 1rem 1.5rem; cursor: pointer;
  transition: opacity 0.2s; -webkit-tap-highlight-color: transparent;
}
.amen-invitation-send:hover { opacity: 0.85; }
.amen-inv-confirmed {
  display: none; font-family: 'Libre Baskerville', serif;
  font-size: 0.95rem; font-style: italic; color: var(--dim);
  line-height: 1.75; text-align: center;
}
.amen-inv-confirmed.amen-inv-confirmed-visible { display: block; }

/* ── Post-liturgy layer (notif + share — rises 3s after commission animates in) ── */
.amen-post-liturgy {
  width: 100%; text-align: center;
  opacity: 0; transform: translateY(-4px);
  max-height: 0; overflow: hidden; margin-bottom: 0;
  transition: opacity 0.9s ease, transform 0.9s ease,
              max-height 0.9s ease, margin-bottom 0.9s ease;
  pointer-events: none;
}
.amen-post-liturgy.amen-visible {
  opacity: 1; transform: translateY(0);
  max-height: 1200px; margin-bottom: 1.6rem;
  pointer-events: auto;
}

.amen-notif-section {
  display: none;
  width: 100%;
  text-align: center;
  padding: 1.4rem 0 0;
  margin-bottom: 1rem;
}
.amen-notif-prompt {
  font-family: 'Libre Baskerville', serif;
  font-size: 0.95rem;
  font-style: italic;
  color: var(--dim);
  line-height: 1.85;
  margin-bottom: 1.2rem;
}

/* Amen screen fade overlays */
.palette-eastertide .screen-amen .screen-fade-top
  { background: linear-gradient(to bottom, #FAF7F2 40%, rgba(250,247,242,0)); }
.palette-eastertide .screen-amen .screen-fade-bottom
  { background: linear-gradient(to top, #FAF7F2 50%, rgba(250,247,242,0)); }
.palette-lent .screen-amen .screen-fade-top
  { background: linear-gradient(to bottom, #241E1A 40%, rgba(36,30,26,0)); }
.palette-lent .screen-amen .screen-fade-bottom
  { background: linear-gradient(to top, #241E1A 50%, rgba(36,30,26,0)); }
.palette-goodfriday .screen-amen .screen-fade-top
  { background: linear-gradient(to bottom, #0E0C0B 40%, rgba(14,12,11,0)); }
.palette-goodfriday .screen-amen .screen-fade-bottom
  { background: linear-gradient(to top, #0E0C0B 50%, rgba(14,12,11,0)); }
.palette-pentecost .screen-amen .screen-fade-top
  { background: linear-gradient(to bottom, #8B1A1A 40%, rgba(139,26,26,0)); }
.palette-pentecost .screen-amen .screen-fade-bottom
  { background: linear-gradient(to top, #8B1A1A 50%, rgba(139,26,26,0)); }
.palette-ordinarytime .screen-amen .screen-fade-top
  { background: linear-gradient(to bottom, #FAF7F2 40%, rgba(250,247,242,0)); }
.palette-ordinarytime .screen-amen .screen-fade-bottom
  { background: linear-gradient(to top, #FAF7F2 50%, rgba(250,247,242,0)); }
.palette-ordinarytime-major .screen-amen .screen-fade-top
  { background: linear-gradient(to bottom, #1A2B1A 40%, rgba(26,43,26,0)); }
.palette-ordinarytime-major .screen-amen .screen-fade-bottom
  { background: linear-gradient(to top, #1A2B1A 50%, rgba(26,43,26,0)); }
.palette-white .screen-amen .screen-fade-top,
.palette-white-major .screen-amen .screen-fade-top
  { background: linear-gradient(to bottom, #FEFEFE 40%, rgba(254,254,254,0)); }
.palette-white .screen-amen .screen-fade-bottom,
.palette-white-major .screen-amen .screen-fade-bottom
  { background: linear-gradient(to top, #FEFEFE 50%, rgba(254,254,254,0)); }
.palette-advent .screen-amen .screen-fade-top,
.palette-rose .screen-amen .screen-fade-top
  { background: linear-gradient(to bottom, #FAF7F2 40%, rgba(250,247,242,0)); }
.palette-advent .screen-amen .screen-fade-bottom,
.palette-rose .screen-amen .screen-fade-bottom
  { background: linear-gradient(to top, #FAF7F2 50%, rgba(250,247,242,0)); }
.palette-black .screen-amen .screen-fade-top
  { background: linear-gradient(to bottom, #1A1A1A 40%, rgba(26,26,26,0)); }
.palette-black .screen-amen .screen-fade-bottom
  { background: linear-gradient(to top, #1A1A1A 50%, rgba(26,26,26,0)); }

/* Production top bar — centred */
.top-bar-inner--prod { justify-content: center; }
"""


JS = r"""
const TOTAL_SCREENS = 7;
const LABELS = ["Season", "Word", "Scripture", "Contemplation", "Prayer", "Practice", "Amen"];

let currentDay    = 0;
let currentScreen = 0;
const totalDays   = DAYS.length;

// ── Palette ────────────────────────────────────────────────────────────────

function applyPalette(dayIndex) {
  const app = document.getElementById('app');
  app.className = app.className.replace(/palette-\S+/g, '').trim();
  app.classList.add(DAYS[dayIndex].palette);
}

function updateDayLabel() {
  const el = document.getElementById('dayLabel');
  if (el) el.textContent = DAYS[currentDay].label;
}

// ── Date utilities ─────────────────────────────────────────────────────────

function parseDayDate(dateStr) {
  if (!dateStr) return null;
  const months = {
    'January':'01','February':'02','March':'03','April':'04','May':'05','June':'06',
    'July':'07','August':'08','September':'09','October':'10','November':'11','December':'12'
  };
  const m = dateStr.match(/^(\d{1,2})\s+(\w+)\s+(\d{4})$/);
  if (!m) return null;
  const mon = months[m[2]];
  if (!mon) return null;
  return m[3] + '-' + mon + '-' + m[1].padStart(2,'0');
}

function todayISO() {
  const now = new Date();
  const y = now.getFullYear();
  const m = String(now.getMonth()+1).padStart(2,'0');
  const d = String(now.getDate()).padStart(2,'0');
  return y + '-' + m + '-' + d;
}

function findTodayIndex() {
  const today = todayISO();
  for (let i = 0; i < DAYS.length; i++) {
    if (parseDayDate(DAYS[i].date) === today) return i;
  }
  return -1;
}

// ── Screen navigation ──────────────────────────────────────────────────────

function getCurrentScreens() {
  return document.querySelectorAll(`.day-screens[data-day="${currentDay}"] .screen`);
}

function getScreenWidth() {
  const appEl = document.getElementById('app');
  return appEl ? appEl.offsetWidth : window.innerWidth;
}

let _transitioning = false;

function goTo(i, anim=true) {
  if (_transitioning) return;
  if (i >= TOTAL_SCREENS) { wrapAmenToScreen1(); return; }
  if (i < 0) i = 0;

  const screens = getCurrentScreens();
  const prev    = screens[currentScreen];
  const next    = screens[i];

  currentScreen = i;
  document.querySelectorAll('.dot').forEach((d, idx) => d.classList.toggle('active', idx === i));
  document.getElementById('counter').textContent = `${i + 1} / ${TOTAL_SCREENS}`;
  const hint = document.getElementById('swipeHint');
  if (hint && i > 0) hint.style.opacity = '0';

  if (!anim) {
    screens.forEach(s => s.classList.remove('active', 'leaving', 'animating'));
    next.classList.add('active');
    return;
  }

  if (prev === next) return;

  _transitioning = true;

  prev.style.zIndex = '2';
  next.style.zIndex = '1';
  prev.classList.remove('active');
  prev.classList.add('leaving');
  next.classList.add('active');

  animateScreen(i);

  setTimeout(() => {
    prev.classList.remove('leaving');
    prev.style.zIndex = '';
    next.style.zIndex = '';
    _transitioning = false;
  }, 880);
}

function wrapAmenToScreen1() {
  if (_transitioning) return;
  _transitioning = true;

  const screens = getCurrentScreens();
  const prev    = screens[currentScreen];
  const veil    = document.getElementById('thresholdVeil');

  // Fade dark veil in
  veil.classList.remove('threshold-opening');
  veil.classList.add('threshold-closing');
  veil.style.pointerEvents = 'all';

  // Silently switch to screen 1 while veil is opaque
  setTimeout(() => {
    screens.forEach(s => s.classList.remove('active', 'leaving', 'animating'));
    currentScreen = 0;
    screens[0].classList.add('active');
    const inner = screens[0].querySelector('.screen-inner');
    if (inner) inner.scrollTop = 0;
    document.querySelectorAll('.dot').forEach((d, idx) => d.classList.toggle('active', idx === 0));
    document.getElementById('counter').textContent = `1 / ${TOTAL_SCREENS}`;
  }, 650);

  // Lift the veil, revealing screen 1
  setTimeout(() => {
    veil.classList.remove('threshold-closing');
    veil.classList.add('threshold-opening');
    animateScreen(0);
  }, 750);

  // Full reset
  setTimeout(() => {
    veil.classList.remove('threshold-opening', 'threshold-closing');
    veil.style.pointerEvents = 'none';
    _transitioning = false;
  }, 1550);
}

function animateScreen(i) {
  const screens = getCurrentScreens();
  const sc = screens[i];
  if (!sc) return;
  const inner = sc.querySelector('.screen-inner');
  if (inner) inner.scrollTop = 0;
  sc.classList.remove('animating');
  void sc.offsetWidth;
  sc.classList.add('animating');
  sc.addEventListener('animationend', () => sc.classList.remove('animating'), {once: true});
}

function buildDots() {
  const el = document.getElementById('dots');
  el.innerHTML = LABELS.map((l, i) =>
    `<button class="dot${i === 0 ? ' active' : ''}" data-index="${i}" aria-label="${l}"></button>`
  ).join('');
  el.querySelectorAll('.dot').forEach(d =>
    d.addEventListener('click', () => goTo(parseInt(d.dataset.index)))
  );
}

function showDay(dayIndex, screenIndex=0) {
  document.querySelectorAll('.day-screens').forEach(el => el.style.display = 'none');
  const dayEl = document.querySelector(`.day-screens[data-day="${dayIndex}"]`);
  if (dayEl) dayEl.style.display = '';
  currentDay    = dayIndex;
  currentScreen = screenIndex;
  applyPalette(dayIndex);
  buildDots();
  updateDayLabel();
  goTo(screenIndex, false);
}

// ── Figure modal ───────────────────────────────────────────────────────────

function openFigure(key) {
  const dayFigs = FIGURES_BY_DAY[currentDay] || {};
  const fig = dayFigs[key]; if (!fig) return;
  document.getElementById('figureName').textContent  = fig.name;
  document.getElementById('figureDates').textContent = fig.dates;
  const roleEl = document.getElementById('figureRole');
  roleEl.textContent = fig.role || '';
  roleEl.style.display = fig.role ? '' : 'none';
  document.getElementById('figureBody').textContent  = fig.body;
  const overlay = document.getElementById('figureOverlay');
  const card = overlay.querySelector('.figure-card');
  // Start card off-position, then animate in after overlay fades up.
  card.style.transition = 'none';
  card.style.opacity = '0';
  card.style.transform = 'translateY(10px)';
  void overlay.offsetHeight;
  overlay.classList.add('open');
  // Let the overlay begin fading, then animate the card in.
  setTimeout(() => {{
    card.style.transition = 'opacity 0.75s ease, transform 0.75s cubic-bezier(0.22,1,0.36,1)';
    card.style.opacity = '1';
    card.style.transform = 'translateY(0)';
  }}, 50);
}
function closeFigure() {{
  const ov = document.getElementById('figureOverlay');
  ov.style.opacity = '0';
  setTimeout(() => {{
    ov.classList.remove('open');
    ov.style.opacity = '';
    const card = ov.querySelector('.figure-card');
    card.style.transition = 'none';
    card.style.opacity = '0';
    card.style.transform = 'translateY(10px)';
  }}, 750);
}}
document.getElementById('figureClose').addEventListener('click', closeFigure);
document.getElementById('figureOverlay').addEventListener('click', ev => {
  if (ev.target === document.getElementById('figureOverlay')) closeFigure();
});
document.addEventListener('click', ev => {
  const btn = ev.target.closest('.figure-link');
  if (btn) { ev.stopPropagation(); openFigure(btn.dataset.figure); }
});

// ── Touch / mouse swipe ────────────────────────────────────────────────────

let tx=0, ty=0, ta=false;
document.addEventListener('touchstart', ev => {
  if (document.getElementById('figureOverlay').classList.contains('open')) return;
  if (ev.target.closest('.day-nav')) return;
  if (ev.target.closest('.figure-link')) return;
  tx = ev.touches[0].clientX; ty = ev.touches[0].clientY; ta = true;
}, {passive: true});
document.addEventListener('touchend', ev => {
  if (!ta) return; ta = false;
  const dx = ev.changedTouches[0].clientX - tx;
  const dy = ev.changedTouches[0].clientY - ty;
  if (Math.abs(dx) > Math.abs(dy) && Math.abs(dx) > 40)
    goTo(currentScreen + (dx < 0 ? 1 : -1));
}, {passive: true});

let mx=0, md=false;
document.addEventListener('mousedown', ev => {
  if (ev.target.closest('.figure-link,.figure-overlay,.dot,.close,.screen-inner,.day-nav')) return;
  mx = ev.clientX; md = true;
});
document.addEventListener('mouseup', ev => {
  if (!md) return; md = false;
  const dx = ev.clientX - mx;
  if (Math.abs(dx) > 55) goTo(currentScreen + (dx < 0 ? 1 : -1));
});

document.addEventListener('keydown', ev => {
  if (document.getElementById('figureOverlay').classList.contains('open')) {
    if (ev.key === 'Escape') closeFigure();
    return;
  }
  if (ev.key === 'ArrowRight' || ev.key === 'ArrowDown') goTo(currentScreen + 1);
  if (ev.key === 'ArrowLeft'  || ev.key === 'ArrowUp')   goTo(currentScreen - 1);
});

// ── Word title auto-fit ────────────────────────────────────────────────────

function fitWordTitle(dayIndex) {
  const dayEl = document.querySelector(`.day-screens[data-day="${dayIndex}"]`);
  if (!dayEl) return;
  dayEl.querySelectorAll('.word-title').forEach(el => {
    el.style.fontSize = '';
    const container = el.closest('.screen-inner') || el.parentElement;
    if (!container) return;
    const style = getComputedStyle(container);
    const usable = container.clientWidth - parseFloat(style.paddingLeft) - parseFloat(style.paddingRight);
    let lo = 14, hi = 128;
    for (let iter = 0; iter < 14; iter++) {
      const mid = (lo + hi) / 2;
      el.style.fontSize = mid + 'px';
      if (el.scrollWidth <= usable) lo = mid; else hi = mid;
    }
    el.style.fontSize = Math.floor(lo) + 'px';
  });
}

function fitAmenWordName(dayIndex) {
  const dayEl = document.querySelector(`.day-screens[data-day="${dayIndex}"]`);
  if (!dayEl) return;
  dayEl.querySelectorAll('.amen-word-name').forEach(el => {
    el.style.fontSize = '';
    const container = el.closest('.screen-inner') || el.parentElement;
    if (!container) return;
    const style = getComputedStyle(container);
    const usable = container.clientWidth - parseFloat(style.paddingLeft) - parseFloat(style.paddingRight);
    // Cap at 3.6rem (58px at 16px base) — match the clamp maximum
    const maxPx = parseFloat(getComputedStyle(document.documentElement).fontSize) * 3.6;
    let lo = 14, hi = Math.min(maxPx, 128);
    for (let iter = 0; iter < 14; iter++) {
      const mid = (lo + hi) / 2;
      el.style.fontSize = mid + 'px';
      if (el.scrollWidth <= usable) lo = mid; else hi = mid;
    }
    el.style.fontSize = Math.floor(lo) + 'px';
  });
}

// ── Desktop navigation arrows ──────────────────────────────────────────────

const desktopPrev = document.getElementById('desktopPrev');
const desktopNext = document.getElementById('desktopNext');

function updateDesktopArrows() {
  if (!desktopPrev || !desktopNext) return;
  desktopPrev.classList.toggle('disabled', currentScreen <= 0);
  desktopNext.classList.toggle('disabled', currentScreen >= TOTAL_SCREENS - 1);
}

if (desktopPrev) desktopPrev.addEventListener('click', () => {
  if (currentScreen > 0) goTo(currentScreen - 1);
});
if (desktopNext) desktopNext.addEventListener('click', () => {
  if (currentScreen < TOTAL_SCREENS - 1) goTo(currentScreen + 1);
  else wrapAmenToScreen1();
});

// ── Scroll overflow indicator ──────────────────────────────────────────────

function getFadeHeight(sc) {
  const fade = sc.querySelector('.screen-fade-bottom');
  return fade ? fade.offsetHeight : 0;
}
function getContentBottom(inner) {
  const children = Array.from(inner.children);
  if (!children.length) return 0;
  const last = children[children.length - 1];
  return last.offsetTop + last.offsetHeight;
}
function hasHiddenContent(inner, sc) {
  const fadeH = getFadeHeight(sc);
  const contentBottom = getContentBottom(inner);
  const viewBottom = inner.scrollTop + inner.clientHeight;
  return contentBottom > viewBottom - fadeH + 4;
}
function scrollElementClearFade(inner, sc, el, behavior) {
  behavior = behavior || 'smooth';
  const fadeH = getFadeHeight(sc);
  const elBottom = el.offsetTop + el.offsetHeight;
  const targetScrollTop = elBottom - inner.clientHeight + fadeH + 16;
  if (targetScrollTop > inner.scrollTop) inner.scrollTo({top: targetScrollTop, behavior});
}

function updateScrollIndicator(screenIndex) {
  const screens = getCurrentScreens();
  const sc = screens[screenIndex];
  if (!sc) return;
  const inner = sc.querySelector('.screen-inner');
  if (!inner) return;
  let ind = sc.querySelector('.scroll-indicator');
  if (!ind) {
    ind = document.createElement('div');
    ind.className = 'scroll-indicator';
    ind.innerHTML = '<svg width="22" height="22" viewBox="0 0 22 22" fill="none"><polyline points="5,8 11,14 17,8" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/></svg>';
    sc.appendChild(ind);
  }
  function check() { ind.classList.toggle('visible', hasHiddenContent(inner, sc)); }
  check();
  inner.removeEventListener('scroll', inner._scrollCheck);
  inner._scrollCheck = check;
  inner.addEventListener('scroll', check, {passive: true});
}

// Hook scroll indicator, desktop arrows and notification prompt into goTo
const _goToOrig = goTo;
goTo = function(i, anim=true) {
  _goToOrig(i, anim);
  updateDesktopArrows();
  const idx = i < 0 ? 0 : i >= TOTAL_SCREENS ? 0 : i;
  setTimeout(() => updateScrollIndicator(idx), 350);
  setTimeout(() => vigilMaybeShowNotifPrompt(idx), 400);
};

// Initialise: fit word titles, update arrows, start scroll indicator
setTimeout(() => {
  for (let d = 0; d < totalDays; d++) { fitWordTitle(d); fitAmenWordName(d); }
  updateDesktopArrows();
  updateScrollIndicator(0);
}, 300);

// Recalculate word titles on resize
window.addEventListener('resize', () => {
  for (let d = 0; d < totalDays; d++) { fitWordTitle(d); fitAmenWordName(d); }
});

// ── Screen 7: Amen ─────────────────────────────────────────────────────────

const VIGIL_COUNTER_URL     = null;
const VIGIL_COUNT_THRESHOLD = 500;

(async function fetchAmenCount() {
  if (!VIGIL_COUNTER_URL) return;
  try {
    const res  = await fetch(VIGIL_COUNTER_URL);
    const data = await res.json();
    if (data && typeof data.count === 'number' && data.count >= VIGIL_COUNT_THRESHOLD) {
      document.querySelectorAll('[id^="amen-count-num-"]').forEach(el => {
        el.textContent = data.count.toLocaleString('en-GB');
      });
      document.querySelectorAll('[id^="amen-count-"]').forEach(el => {
        if (el.id.startsWith('amen-count-b') || el.classList.contains('amen-count-block'))
          el.classList.add('amen-count-visible');
      });
    }
  } catch(e) {}
})();

async function amenIncrementCount() {
  if (!VIGIL_COUNTER_URL) return;
  try { await fetch(VIGIL_COUNTER_URL + '/increment', { method: 'POST' }); } catch(e) {}
}

const amenCounted = {};

function amenCheck(row) {
  const uid     = row.dataset.uid;
  const box     = document.getElementById('amen-box-' + uid);
  const hero    = document.getElementById('amen-hero-' + uid);
  const word    = document.getElementById('amen-word-' + uid);
  const comm    = document.getElementById('amen-commission-' + uid);
  const post    = document.getElementById('amen-post-' + uid);
  const share   = document.getElementById('amen-share-' + uid);
  const checked = box.classList.toggle('amen-checked');
  row.setAttribute('aria-checked', String(checked));
  if (checked) {
    if (!amenCounted[uid]) { amenCounted[uid] = true; amenIncrementCount(); }
    vigilRecordKept();
    amenPrebuildShareImage(uid);

    // Update completion count before deciding what to show
    let count = parseInt(localStorage.getItem('vigil-completion-count') || '0', 10);
    count = Math.max(0, count) + 1;
    localStorage.setItem('vigil-completion-count', String(count));

    setTimeout(() => { if (hero) hero.classList.add('amen-hero-checked'); }, 0);
    setTimeout(() => word.classList.add('amen-visible'), 300);
    // Commission appears 0.7s after tick; its own transition takes 0.9s
    setTimeout(() => comm.classList.add('amen-visible'), 700);
    // Post-liturgy layer rises 3s after commission animation completes: 700 + 900 + 3000
    setTimeout(() => {
      amenShowPostLiturgy(uid, count);
    }, 4600);
  } else {
    // Decrement count (min 0)
    let count = parseInt(localStorage.getItem('vigil-completion-count') || '0', 10);
    count = Math.max(0, count - 1);
    localStorage.setItem('vigil-completion-count', String(count));

    if (post)  post.classList.remove('amen-visible');
    setTimeout(() => comm.classList.remove('amen-visible'), 300);
    setTimeout(() => word.classList.remove('amen-visible'), 600);
    setTimeout(() => { if (hero) hero.classList.remove('amen-hero-checked'); }, 900);
    const initial   = document.getElementById('amen-share-initial-' + uid);
    const confirmed = document.getElementById('amen-share-confirmed-' + uid);
    if (initial)   initial.style.display = '';
    if (confirmed) { confirmed.textContent = ''; confirmed.classList.remove('amen-share-confirmed-visible'); }
  }
}

function amenShowPostLiturgy(uid, completionCount) {
  const post  = document.getElementById('amen-post-' + uid);
  const share = document.getElementById('amen-share-' + uid);
  if (!post) return;

  // Show the notif section if applicable
  const notifShown = amenMaybeShowNotif(uid);

  // Show the share ask if: count ≥ 2 AND share not permanently dismissed
  const shareDismissed = localStorage.getItem('vigil-share-dismissed');
  const showShare = !shareDismissed && completionCount >= 2;

  if (showShare && share) {
    // If notif was shown, delay share by 1.5s after the user resolves it;
    // since we can't know when that is, we use a conservative minimum delay.
    // The notif section hides itself when dismissed; share appears 1.5s later.
    if (notifShown) {
      // Mark that share should appear after notif resolves
      share.dataset.pendingAfterNotif = '1';
    } else {
      // No notif — show share immediately as post-liturgy layer rises
      share.classList.add('amen-visible');
    }
  }

  post.classList.add('amen-visible');
}

// ── Share image: pre-render on checkbox tick, share synchronously on button tap ──

const amenShareBlobs = new Map();

async function amenPrebuildShareImage(uid) {
  // Pure Canvas 2D rendering — no SVG foreignObject, works on iOS Safari.
  // ctx.letterSpacing is NOT used anywhere — it is unsupported in WebKit webviews
  // and causes font fallback failures. Spacing is achieved via pre-spaced strings
  // or accepted as-is at small sizes.

  const wordEl   = document.getElementById('amen-word-' + uid);
  const wordName = wordEl ? (wordEl.querySelector('.amen-word-name')?.textContent?.trim() || '') : '';
  // Read palette colours from the app element, which carries the palette class.
  // documentElement does not have the palette class and would return fallbacks.
  const appEl = document.getElementById('app') || document.documentElement;
  const cs   = getComputedStyle(appEl);
  const BG   = cs.getPropertyValue('--bg').trim()   || '#FAF7F2';
  const GOLD = cs.getPropertyValue('--gold').trim() || '#B8860B';
  const TEXT = cs.getPropertyValue('--text').trim() || '#1C1714';
  const DIM  = cs.getPropertyValue('--dim').trim()  || '#8A7D6E';

  // ── Guarantee fonts are loaded before drawing ──────────────────────────────
  // document.fonts.ready is not enough on all platforms — we explicitly load
  // each required face via FontFace so the canvas engine has them available.
  const GFONTS_URL = 'https://fonts.googleapis.com/css2?family=Cormorant+Garamond:ital,wght@0,300;0,600;1,300;1,400&family=Raleway:wght@300;400&display=swap';
  async function ensureFonts() {
    // Step 1: wait for any already-loading fonts
    if (document.fonts && document.fonts.ready) await document.fonts.ready;
    // Step 2: check if our key faces are present; if not, load via FontFace API
    const needed = [
      { family: 'Cormorant Garamond', weight: '600', style: 'normal' },
      { family: 'Cormorant Garamond', weight: '300', style: 'italic' },
      { family: 'Cormorant Garamond', weight: '400', style: 'italic' },
      { family: 'Raleway',            weight: '300', style: 'normal' },
      { family: 'Raleway',            weight: '400', style: 'normal' },
    ];
    const missing = needed.filter(f => !document.fonts.check(`${f.style} ${f.weight} 16px "${f.family}"`));
    if (missing.length > 0) {
      // Inject a temporary hidden element to force the Google Fonts stylesheet
      // to load all needed weights, then wait for fonts.ready again.
      const probe = document.createElement('span');
      probe.style.cssText = 'position:fixed;left:-9999px;opacity:0;font-size:4px;';
      probe.innerHTML =
        '<span style="font-family:\'Cormorant Garamond\';font-weight:600;">A</span>' +
        '<span style="font-family:\'Cormorant Garamond\';font-weight:300;font-style:italic;">A</span>' +
        '<span style="font-family:\'Cormorant Garamond\';font-weight:400;font-style:italic;">A</span>' +
        '<span style="font-family:\'Raleway\';font-weight:300;">A</span>' +
        '<span style="font-family:\'Raleway\';font-weight:400;">A</span>';
      document.body.appendChild(probe);
      await document.fonts.ready;
      // Give WebKit an extra tick to finish rasterisation
      await new Promise(r => setTimeout(r, 80));
      document.body.removeChild(probe);
    }
  }
  await ensureFonts();

  // ── Helpers ────────────────────────────────────────────────────────────────
  function hexToRgba(hex, alpha) {
    const r = parseInt(hex.slice(1,3),16);
    const g = parseInt(hex.slice(3,5),16);
    const b = parseInt(hex.slice(5,7),16);
    return `rgba(${r},${g},${b},${alpha})`;
  }
  function goldRgba(a) { return GOLD.startsWith('#') ? hexToRgba(GOLD, a) : `rgba(184,134,11,${a})`; }

  // ── Content values ─────────────────────────────────────────────────────────
  // Date: uppercase, e.g. "1 MAY 2026"
  const dayLabel  = (DAYS[currentDay] && DAYS[currentDay].label) ? DAYS[currentDay].label : '';
  const dateUpper = dayLabel.toUpperCase();
  const wordUpper = wordName.toUpperCase();

  // ── Canvas dimensions ──────────────────────────────────────────────────────
  const SCALE = 3;  // 3× for crisp rendering on Retina / high-DPI screens
  const W     = 375; // logical card width in px

  // ── Find the largest word font size that fits the card width ───────────────
  const measureC   = document.createElement('canvas');
  const measureCtx = measureC.getContext('2d');
  let wSize = 88;
  measureCtx.font = `600 ${wSize}px "Cormorant Garamond", Georgia, serif`;
  while (measureCtx.measureText(wordUpper).width > W - 48 && wSize > 32) {
    wSize -= 2;
    measureCtx.font = `600 ${wSize}px "Cormorant Garamond", Georgia, serif`;
  }

  // ── Layout constants (all in logical px) ──────────────────────────────────
  const topPad       = 48;
  const vigilSize    = 22;   // "V I G I L" — spaces give the letter-spacing feel
  const dateSize     = 11;
  const checkboxSz   = 24;   // checkbox square
  const checkLabelSz = 20;   // "I kept Vigil today."
  const shortRuleW   = 52;
  const commSize     = 19;
  const commLH       = 1.75;
  const urlSize      = 11;
  const btmPad       = 44;

  const gapVigilDate   = 10;
  const gapDateCheck   = 30;
  const gapCheckRule   = 24;
  const gapRuleWord    = 20;
  const gapWordTopRule = 28;
  const gapRuleComm    = 22;
  const gapCommBotRule = 26;
  const gapBotRuleUrl  = 22;

  const commLineH = commSize * commLH;

  // Measure checkbox label width using the actual font (post-load)
  measureCtx.font = `300 italic ${checkLabelSz}px "Cormorant Garamond", Georgia, serif`;
  const labelW = measureCtx.measureText('I kept Vigil today.').width;
  const rowW   = checkboxSz + 14 + labelW;
  const rowX   = (W - rowW) / 2;

  const H = Math.ceil(
    topPad +
    vigilSize    + gapVigilDate   +
    dateSize     + gapDateCheck   +
    checkboxSz   + gapCheckRule   +
    1            + gapRuleWord    +
    wSize        + gapWordTopRule +
    1            + gapRuleComm    +
    commLineH    + commLineH      + gapCommBotRule +
    1            + gapBotRuleUrl  +
    urlSize      + btmPad
  );

  // ── Create canvas ──────────────────────────────────────────────────────────
  const c   = document.createElement('canvas');
  c.width   = W * SCALE;
  c.height  = H * SCALE;
  const ctx = c.getContext('2d');
  ctx.scale(SCALE, SCALE);

  // Background
  ctx.fillStyle = BG;
  ctx.fillRect(0, 0, W, H);

  let y = topPad;

  // ── "V I G I L" ───────────────────────────────────────────────────────────
  // Spaces between letters give the tracking feel without ctx.letterSpacing
  ctx.fillStyle    = GOLD;
  ctx.font         = `300 ${vigilSize}px "Raleway", "Helvetica Neue", Arial, sans-serif`;
  ctx.textAlign    = 'center';
  ctx.textBaseline = 'top';
  ctx.fillText('V I G I L', W / 2, y);
  y += vigilSize + gapVigilDate;

  // ── Date line ─────────────────────────────────────────────────────────────
  ctx.fillStyle = DIM;
  ctx.font      = `400 ${dateSize}px "Raleway", "Helvetica Neue", Arial, sans-serif`;
  ctx.fillText(dateUpper, W / 2, y);
  y += dateSize + gapDateCheck;

  // ── Checkbox square ────────────────────────────────────────────────────────
  ctx.lineWidth   = 1.5;
  ctx.fillStyle   = goldRgba(0.06);
  ctx.fillRect(rowX, y, checkboxSz, checkboxSz);
  ctx.strokeStyle = GOLD;
  ctx.strokeRect(rowX + 0.75, y + 0.75, checkboxSz - 1.5, checkboxSz - 1.5);

  // Tick mark — polyline scaled to fit the checkbox
  const st = checkboxSz / 13;
  const tx = rowX + (checkboxSz - 13 * st) / 2;
  const ty = y    + (checkboxSz - 13 * st) / 2;
  ctx.beginPath();
  ctx.moveTo(tx + 2  * st, ty + 6.5 * st);
  ctx.lineTo(tx + 5  * st, ty + 10  * st);
  ctx.lineTo(tx + 11 * st, ty + 3   * st);
  ctx.strokeStyle = GOLD;
  ctx.lineWidth   = 1.6;
  ctx.lineCap     = 'round';
  ctx.lineJoin    = 'round';
  ctx.stroke();

  // ── "I kept Vigil today." ─────────────────────────────────────────────────
  ctx.fillStyle    = TEXT;
  ctx.font         = `300 italic ${checkLabelSz}px "Cormorant Garamond", Georgia, serif`;
  ctx.textAlign    = 'left';
  ctx.textBaseline = 'middle';
  ctx.fillText('I kept Vigil today.', rowX + checkboxSz + 14, y + checkboxSz / 2);
  y += checkboxSz + gapCheckRule;

  // ── Short gold rule ────────────────────────────────────────────────────────
  ctx.globalAlpha = 0.5;
  ctx.fillStyle   = GOLD;
  ctx.fillRect((W - shortRuleW) / 2, y, shortRuleW, 1);
  ctx.globalAlpha = 1;
  y += 1 + gapRuleWord;

  // ── Word of the day ────────────────────────────────────────────────────────
  ctx.fillStyle    = TEXT;
  ctx.font         = `600 ${wSize}px "Cormorant Garamond", Georgia, serif`;
  ctx.textAlign    = 'center';
  ctx.textBaseline = 'top';
  ctx.fillText(wordUpper, W / 2, y);
  y += wSize + gapWordTopRule;

  // ── Full-width rule above commission (gold 25%) ────────────────────────────
  ctx.globalAlpha = 0.25;
  ctx.fillStyle   = GOLD;
  ctx.fillRect(0, y, W, 1);
  ctx.globalAlpha = 1;
  y += 1 + gapRuleComm;

  // ── Commission phrase ──────────────────────────────────────────────────────
  ctx.fillStyle    = TEXT;
  ctx.font         = `400 italic ${commSize}px "Cormorant Garamond", Georgia, serif`;
  ctx.textAlign    = 'center';
  ctx.textBaseline = 'top';
  ctx.fillText('Go in peace.', W / 2, y);
  y += commLineH;
  ctx.fillText('And carry the Word with you.', W / 2, y);
  y += commLineH + gapCommBotRule;

  // ── Full-width rule below commission (gold 25%) ────────────────────────────
  ctx.globalAlpha = 0.25;
  ctx.fillStyle   = GOLD;
  ctx.fillRect(0, y, W, 1);
  ctx.globalAlpha = 1;
  y += 1 + gapBotRuleUrl;

  // ── URL ───────────────────────────────────────────────────────────────────
  ctx.fillStyle    = GOLD;
  ctx.font         = `400 ${urlSize}px "Raleway", "Helvetica Neue", Arial, sans-serif`;
  ctx.textAlign    = 'center';
  ctx.textBaseline = 'top';
  ctx.fillText('https://dailyvigil.com', W / 2, y);

  // ── Export ────────────────────────────────────────────────────────────────
  try {
    const blob = await new Promise((resolve, reject) => {
      c.toBlob(b => b ? resolve(b) : reject(new Error('toBlob returned null')), 'image/png');
    });
    amenShareBlobs.set(uid, blob);
  } catch(e) {
    // Prebuild failed silently — amenShareImage will handle fallback
  }
}

async function amenShareImage(uid) {
  const blob = amenShareBlobs.get(uid);
  if (blob) {
    const file = new File([blob], 'vigil-today.png', { type: 'image/png' });
    try {
      if (navigator.canShare && navigator.canShare({ files: [file] })) {
        await navigator.share({ files: [file], title: 'Vigil' });
        amenShowShareConfirmed(uid, 'Thank you for sharing.');
      } else {
        const dlUrl = URL.createObjectURL(blob);
        const a = document.createElement('a');
        a.href = dlUrl; a.download = 'vigil-today.png'; a.click();
        setTimeout(() => URL.revokeObjectURL(dlUrl), 2000);
        amenShowShareConfirmed(uid, 'Image saved to your device.');
      }
    } catch(e) {
      if (e.name !== 'AbortError') amenShowShareConfirmed(uid, 'Image saved to your device.');
    }
  } else {
    amenShowShareConfirmed(uid, 'Preparing image\u2026');
    await amenPrebuildShareImage(uid);
    const b = amenShareBlobs.get(uid);
    if (!b) { amenShowShareConfirmed(uid, 'Unable to generate image. Try again.'); return; }
    const dlUrl = URL.createObjectURL(b);
    const a = document.createElement('a');
    a.href = dlUrl; a.download = 'vigil-today.png'; a.click();
    setTimeout(() => URL.revokeObjectURL(dlUrl), 2000);
    amenShowShareConfirmed(uid, 'Image saved to your device.');
  }
}

function amenShareApp(uid) {
  const text = '\u2018I\u2019ve been keeping Vigil \u2014 a daily devotional app rooted in the Christian year. I like it and thought you might too.';
  amenTriggerShare(text, 'https://dailyvigil.app',
    () => amenShowShareConfirmed(uid, 'Thank you for sharing Vigil.'),
    () => amenShowShareConfirmed(uid, 'The message has been copied. Paste it into a text, email or WhatsApp to share.')
  );
}

function amenShowShareConfirmed(uid, msg) {
  const initial   = document.getElementById('amen-share-initial-' + uid);
  const confirmed = document.getElementById('amen-share-confirmed-' + uid);
  if (initial)   initial.style.display = 'none';
  if (confirmed) { confirmed.innerHTML = msg; confirmed.classList.add('amen-share-confirmed-visible'); }
}

function amenNotNow(uid) {
  const share = document.getElementById('amen-share-' + uid);
  if (!share) return;
  share.style.transition = 'opacity 0.3s ease';
  share.style.opacity    = '0';
  setTimeout(() => {
    share.classList.remove('amen-visible');
    share.style.opacity = '';
  }, 320);
}

const amenInvOpen = {};

function amenToggleInvitation(uid) {
  amenInvOpen[uid] = !amenInvOpen[uid];
  const panel = document.getElementById('amen-inv-panel-' + uid);
  const arrow = document.getElementById('amen-inv-arrow-' + uid);
  if (panel) panel.classList.toggle('amen-inv-open', amenInvOpen[uid]);
  if (arrow) arrow.classList.toggle('amen-inv-arrow-open', amenInvOpen[uid]);
  if (amenInvOpen[uid]) {
    setTimeout(() => {
      const inner = panel ? panel.closest('.screen-inner') : null;
      if (inner) inner.scrollTo({ top: inner.scrollHeight, behavior: 'smooth' });
    }, 480);
  }
}

function amenSendInvitation(uid) {
  const text = "I\u2019ve been using Vigil for my daily devotion \u2014 a beautiful app rooted in the Christian year. I thought you might like it.";
  amenTriggerShare(text, 'https://dailyvigil.app',
    () => amenShowInvConfirmed(uid, 'Invitation sent. Thank you for sharing Vigil.'),
    () => amenShowInvConfirmed(uid, 'Copied to clipboard \u2014 ready to send.')
  );
}

function amenShowInvConfirmed(uid, msg) {
  const actions   = document.getElementById('amen-inv-actions-' + uid);
  const confirmed = document.getElementById('amen-inv-confirmed-' + uid);
  if (actions)   actions.style.display = 'none';
  if (confirmed) { confirmed.textContent = msg; confirmed.classList.add('amen-inv-confirmed-visible'); }
}

function amenTriggerShare(text, url, onSuccess, onCopy) {
  if (typeof navigator.share === 'function') {
    navigator.share({ title: 'Vigil', text, url })
      .then(onSuccess)
      .catch(err => { if (err.name !== 'AbortError') amenCopyText(text + '\n\n' + url, onCopy); });
  } else {
    amenCopyText(text + '\n\n' + url, onCopy);
  }
}

function amenCopyText(text, cb) {
  if (navigator.clipboard && navigator.clipboard.writeText) {
    navigator.clipboard.writeText(text).then(cb).catch(() => amenExecCopy(text, cb));
  } else {
    amenExecCopy(text, cb);
  }
}

function amenExecCopy(text, cb) {
  const ta = document.createElement('textarea');
  ta.value = text; ta.style.cssText = 'position:fixed;opacity:0;top:0;left:0;';
  document.body.appendChild(ta); ta.focus(); ta.select();
  try { document.execCommand('copy'); } catch(e) {}
  document.body.removeChild(ta); if (cb) cb();
}

// ── Notification permission ────────────────────────────────────────────────

// Number of consecutive days kept before the prompt re-appears after dismissal.
const NOTIF_REPROMPT_DAYS = 4;

function vigilTodayISO() {
  const now = new Date();
  return now.getFullYear() + '-'
    + String(now.getMonth() + 1).padStart(2, '0') + '-'
    + String(now.getDate()).padStart(2, '0');
}

// Add today's date to the kept-dates log and trim the list to the last
// NOTIF_REPROMPT_DAYS entries — we never need more than that.
function vigilRecordKept() {
  const today = vigilTodayISO();
  let dates = [];
  try { dates = JSON.parse(localStorage.getItem('vigil-kept-dates') || '[]'); } catch(e) {}
  if (!dates.includes(today)) dates.push(today);
  dates.sort();
  dates = dates.slice(-NOTIF_REPROMPT_DAYS);
  localStorage.setItem('vigil-kept-dates', JSON.stringify(dates));
}

// Returns true if the kept-dates log contains a run of NOTIF_REPROMPT_DAYS
// consecutive calendar days ending on today.
function vigilHasStreak() {
  const today = vigilTodayISO();
  let dates = [];
  try { dates = JSON.parse(localStorage.getItem('vigil-kept-dates') || '[]'); } catch(e) {}
  if (dates.length < NOTIF_REPROMPT_DAYS) return false;
  const recent = dates.slice(-NOTIF_REPROMPT_DAYS);
  // The last entry must be today.
  if (recent[recent.length - 1] !== today) return false;
  // Each entry must be exactly one calendar day before the next.
  for (let i = 1; i < recent.length; i++) {
    const diff = (new Date(recent[i]) - new Date(recent[i - 1])) / 86400000;
    if (diff !== 1) return false;
  }
  return true;
}

function vigilShouldShowNotifPrompt() {
  if (typeof Notification !== 'undefined' && Notification.permission === 'granted') return false;
  if (typeof Notification !== 'undefined' && Notification.permission === 'denied') return false;
  const dismissed = localStorage.getItem('vigil-notif-dismissed');
  // Never shown before — show it.
  if (!dismissed) return true;
  // Permanently dismissed after granting (should be caught above, but guard anyway).
  if (dismissed === 'granted') return false;
  // Re-prompt only when the user has kept a streak of NOTIF_REPROMPT_DAYS consecutive days.
  return vigilHasStreak();
}

// ── Notification browser detection ───────────────────────────────────────────

function vigilGetNotifState() {
  // Returns 'normal', 'ios-safari', or 'ios-other'.
  const ua = navigator.userAgent;
  const isIOS = /iP(hone|ad|od)/.test(ua) ||
                (navigator.platform === 'MacIntel' && navigator.maxTouchPoints > 1);
  if (!isIOS) return 'normal';
  // Installed PWA on iOS behaves like normal — notifications are supported.
  if (window.navigator.standalone === true) return 'normal';
  // Other iOS browsers all inject their own token into the UA string.
  // If none of those tokens are present, we are in Safari.
  const isOtherBrowser = /CriOS|FxiOS|EdgiOS|OPiOS|DuckDuckGo|Brave|SamsungBrowser|MiuiBrowser/.test(ua);
  return isOtherBrowser ? 'ios-other' : 'ios-safari';
}

function vigilMaybeShowNotifPrompt(screenIndex) {
  // No longer shows on screen arrival — the post-liturgy layer handles this.
  // Kept as a no-op stub so existing goTo hook calls do not error.
}

function amenMaybeShowNotif(uid) {
  // Shows the notif section inside the post-liturgy layer if applicable.
  // Returns true if it was shown, false if suppressed.
  if (!vigilShouldShowNotifPrompt()) return false;
  const outer = document.getElementById('amen-notif-' + uid);
  if (!outer) return false;
  const state = vigilGetNotifState();
  const inner = document.getElementById('amen-notif-' + state + '-' + uid);
  if (inner) {
    outer.style.display = 'block';
    inner.style.display = 'block';
    return true;
  }
  return false;
}

function amenTriggerDeferredShare(uid) {
  // Called 1.5s after the notif section resolves, to show the share ask if pending.
  const share = document.getElementById('amen-share-' + uid);
  if (share && share.dataset.pendingAfterNotif) {
    delete share.dataset.pendingAfterNotif;
    const shareDismissed = localStorage.getItem('vigil-share-dismissed');
    if (!shareDismissed) {
      setTimeout(() => {
        share.classList.add('amen-visible');
      }, 1500);
    }
  }
}

function vigilRequestNotification(uid) {
  function doRequest() {
    OneSignal.Notifications.requestPermission().then(function() {
      localStorage.setItem('vigil-notif-dismissed', 'granted');
      const el = document.getElementById('amen-notif-' + uid);
      if (el) el.style.display = 'none';
      amenTriggerDeferredShare(uid);
    });
  }
  // By the time the user taps the button, the SDK will already be initialised.
  // Call it directly. Fall back to the deferred queue only if somehow it is not.
  if (window.OneSignal && window.OneSignal.Notifications) {
    doRequest();
  } else if (window.OneSignalDeferred) {
    OneSignalDeferred.push(async function(OneSignal) { doRequest(); });
  }
}

function vigilDismissNotification(uid) {
  // Record the dismissal date so we can measure streaks from this point,
  // but do not clear the kept-dates log — the streak continues to accrue.
  localStorage.setItem('vigil-notif-dismissed', vigilTodayISO());
  const el = document.getElementById('amen-notif-' + uid);
  if (el) el.style.display = 'none';
  amenTriggerDeferredShare(uid);
}

// ── First-visit welcome overlay ───────────────────────────────────────────

(function initWelcomeOverlay() {
  const overlay = document.getElementById('welcomeOverlay');
  const btn     = document.getElementById('welcomeBegin');
  if (!overlay || !btn) return;

  if (!localStorage.getItem('vigil-welcomed')) {
    overlay.style.display = 'flex';
    btn.addEventListener('click', function() {
      overlay.classList.add('welcome-hiding');
      setTimeout(function() {
        overlay.style.display = 'none';
      }, 320);
      localStorage.setItem('vigil-welcomed', '1');
    });
  }
})();

// ── Desktop notification nudge in welcome overlay ────────────────────────
// On desktop browsers only, adds a short sentence to the first-visit overlay
// reminding the user that notifications can be turned on from the final screen.
// Only shown if the user has not already granted or denied notification permission.

(function addDesktopNotifNudge() {
  // Desktop = not iOS, not Android (no touch-primary device).
  const ua = navigator.userAgent;
  const isIOS     = /iP(hone|ad|od)/.test(ua) ||
                    (navigator.platform === 'MacIntel' && navigator.maxTouchPoints > 1);
  const isAndroid = /Android/.test(ua);
  if (isIOS || isAndroid) return;

  // Only show if notification permission is not yet decided.
  if (typeof Notification !== 'undefined' &&
      (Notification.permission === 'granted' || Notification.permission === 'denied')) return;

  // Inject the nudge sentence before the Begin button.
  const btn = document.getElementById('welcomeBegin');
  if (!btn) return;
  const nudge = document.createElement('p');
  nudge.style.cssText = [
    'font-family:\'Libre Baskerville\',serif',
    'font-size:0.82rem',
    'line-height:1.8',
    'color:var(--dim)',
    'margin:0 0 1.6rem 0',
    'font-style:italic',
  ].join(';');
  nudge.textContent = 'You can also turn on morning notifications from the final screen.';
  btn.parentNode.insertBefore(nudge, btn);
})();

// ── PWA first-launch overlay ──────────────────────────────────────────────
// Shown once when the app is opened from the home screen icon for the first
// time, reminding the user to complete the liturgy and turn on notifications.

(function initPwaWelcomeOverlay() {
  // Only fires when running as an installed PWA (standalone mode).
  if (window.navigator.standalone !== true) return;
  const overlay = document.getElementById('pwaWelcomeOverlay');
  const btn     = document.getElementById('pwaWelcomeBegin');
  if (!overlay || !btn) return;

  if (!localStorage.getItem('vigil-pwa-welcomed')) {
    overlay.style.display = 'flex';
    btn.addEventListener('click', function() {
      overlay.classList.add('welcome-hiding');
      setTimeout(function() {
        overlay.style.display = 'none';
      }, 320);
      localStorage.setItem('vigil-pwa-welcomed', '1');
    });
  }
})();

// ── Production mode: today only ───────────────────────────────────────────

function showNoContentMessage(msg) {
  const app = document.getElementById('app');
  app.innerHTML = `
    <div style="
      display:flex; flex-direction:column; align-items:center; justify-content:center;
      height:100svh; padding:2rem; text-align:center; font-family:'Raleway',sans-serif;
      color:var(--dim);">
      <div style="font-family:'Cormorant Garamond',serif; font-size:2.6rem; font-weight:300;
        letter-spacing:0.4em; text-transform:uppercase; color:var(--gold); margin-bottom:2rem;">
        Vigil
      </div>
      <div style="font-size:1.2rem; line-height:2; max-width:320px;">${msg}</div>
    </div>`;
}

// ── Date override via ?date=YYYY-MM-DD query parameter ────────────────────
// Visiting ?date=2026-05-12 shows that day's content regardless of the
// actual date.  Only works if that day's content is embedded in this file.
// Normal visitors never use this parameter, so their experience is unchanged.

function getDateOverride() {
  try {
    const p = new URLSearchParams(window.location.search);
    const v = p.get('date');
    if (v && /^\d{4}-\d{2}-\d{2}$/.test(v)) return v;
  } catch(e) {}
  return null;
}

function findDateIndex(iso) {
  for (let i = 0; i < DAYS.length; i++) {
    if (parseDayDate(DAYS[i].date) === iso) return i;
  }
  return -1;
}

const dateOverride = getDateOverride();
const targetIdx = dateOverride ? findDateIndex(dateOverride) : findTodayIndex();

if (targetIdx === -1) {
  if (dateOverride) {
    showNoContentMessage('No content found for ' + dateOverride + '.<br>The date may be outside the embedded range.');
  } else {
    const today = todayISO();
    const firstDate = parseDayDate(DAYS[0] && DAYS[0].date);
    if (firstDate && today < firstDate) {
      showNoContentMessage('Vigil begins on ' + DAYS[0].date + '.<br>Come back then.');
    } else {
      showNoContentMessage('Vigil will return with a new word tomorrow.');
    }
  }
} else {
  showDay(targetIdx, 0);

  // Only schedule the midnight advance for normal (non-override) visits,
  // so the preview page stays on the chosen date.
  if (!dateOverride) {
    function msUntilMidnight() {
      const now  = new Date();
      const next = new Date(now.getFullYear(), now.getMonth(), now.getDate()+1, 0, 0, 0, 0);
      return next - now;
    }
    function scheduleMidnightAdvance() {
      setTimeout(() => { window.location.reload(); }, msUntilMidnight());
    }
    scheduleMidnightAdvance();
  }
}

// ── Install nudge ─────────────────────────────────────────────────────────────
// Shows a persistent footer nudge on every screen until the app is installed.
// On iOS Safari: tapping the link shows a modal with step-by-step instructions.
// On Android Chrome: tapping the link triggers the native install prompt.
// Hidden permanently once running in standalone mode (installed PWA).

(function initInstallNudge() {
  const nudge    = document.getElementById('installNudge');
  const link     = document.getElementById('installNudgeLink');
  const overlay  = document.getElementById('installModalOverlay');
  const card     = document.getElementById('installModalCard');
  const closeBtn = document.getElementById('installModalClose');
  const app      = document.getElementById('app');
  if (!nudge || !link || !overlay || !card || !closeBtn) return;

  // Detect standalone mode — if installed, hide nudge permanently.
  const isStandalone = window.navigator.standalone === true ||
                       window.matchMedia('(display-mode: standalone)').matches;
  if (isStandalone) return;

  // Show the nudge.
  nudge.style.display = 'block';

  // Capture Android Chrome install prompt.
  let deferredPrompt = null;
  window.addEventListener('beforeinstallprompt', function(e) {
    e.preventDefault();
    deferredPrompt = e;
  });

  // Detect iOS Safari (to show modal rather than native prompt).
  const ua = navigator.userAgent;
  const isIOS = /iP(hone|ad|od)/.test(ua) ||
                (navigator.platform === 'MacIntel' && navigator.maxTouchPoints > 1);
  const isOtherBrowser = /CriOS|FxiOS|EdgiOS|OPiOS|DuckDuckGo|Brave|SamsungBrowser/.test(ua);
  const isIOSSafari = isIOS && !isOtherBrowser;

  // Animation keyframes.
  const styleEl = document.createElement('style');
  styleEl.textContent = [
    '@keyframes installOverlayIn  { from { opacity:0; } to { opacity:1; } }',
    '@keyframes installOverlayOut { from { opacity:1; } to { opacity:0; } }',
    '@keyframes installCardUp     { from { transform:translateY(100%); opacity:0.6; } to { transform:translateY(0); opacity:1; } }',
    '@keyframes installCardDown   { from { transform:translateY(0); opacity:1; } to { transform:translateY(100%); opacity:0.6; } }',
  ].join('');
  document.head.appendChild(styleEl);

  function openModal() {
    card.style.animation = 'none';
    card.offsetHeight;
    overlay.style.opacity = '0';
    overlay.classList.add('install-modal-visible');
    overlay.offsetHeight;
    overlay.style.animation = 'installOverlayIn 0.32s ease forwards';
    card.style.animation = 'installCardUp 0.55s cubic-bezier(0.22,1,0.36,1) forwards';
    app.classList.add('install-modal-open');
  }

  function closeModal() {
    overlay.style.animation = 'installOverlayOut 0.32s ease forwards';
    card.style.animation = 'installCardDown 0.32s cubic-bezier(0.4,0,1,1) forwards';
    setTimeout(function() {
      overlay.classList.remove('install-modal-visible');
      app.classList.remove('install-modal-open');
    }, 320);
  }

  link.addEventListener('click', function() {
    if (deferredPrompt) {
      // Android Chrome — trigger native prompt directly.
      deferredPrompt.prompt();
      deferredPrompt.userChoice.then(function(result) {
        if (result.outcome === 'accepted') nudge.style.display = 'none';
        deferredPrompt = null;
      });
    } else {
      // iOS Safari and all other browsers — show the modal.
      openModal();
    }
  });

  closeBtn.addEventListener('click', closeModal);

  overlay.addEventListener('click', function(e) {
    if (e.target === overlay) closeModal();
  });

  // Hide nudge permanently once installed.
  window.addEventListener('appinstalled', function() {
    nudge.style.display = 'none';
  });
})();
"""


def build_html(days_window, today_day, build_time):
    """Assemble the complete index.html string.

    days_window — list of (day, figures) tuples, from today through
                  today + DAYS_WINDOW - 1. The JS selects the correct
                  entry by matching the user's local date at runtime.
    today_day   — the day dict for today (used only for the build comment).
    """
    days_js, figures_js = build_days_js(days_window)
    all_days_html = build_all_days_html(days_window)

    # Build comment: show today's word and date range embedded
    raw_word = today_day["word"].strip().split("\n")[0].strip()
    m = re.match(r'^([A-Z\s\-\']+?)\s*(\[|$)', raw_word)
    word_display = m.group(1).strip() if m else raw_word
    d = parse_date_from_liturgical_day(today_day["liturgical_day"])
    date_display = d.strftime("%-d %B %Y") if d else "unknown date"
    last_day, _ = days_window[-1]
    d_last = parse_date_from_liturgical_day(last_day["liturgical_day"])
    last_display = d_last.strftime("%-d %B %Y") if d_last else "?"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<!-- Vigil · Production MVP · Generated {build_time} · {date_display} – {last_display} ({len(days_window)} day(s)) -->
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Vigil</title>
<!-- PWA / icon declarations -->
<link rel="manifest" href="/manifest.json">
<link rel="apple-touch-icon" href="/icons/icon-180.png">
<link rel="icon" type="image/png" sizes="32x32" href="/icons/icon-32.png">
<link rel="icon" type="image/png" sizes="16x16" href="/icons/icon-16.png">
<meta name="theme-color" content="#FAF7F2">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="default">
<meta name="apple-mobile-web-app-title" content="Vigil">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Cormorant+Garamond:ital,wght@0,300;0,400;0,500;0,600;1,300;1,400;1,500&family=Libre+Baskerville:ital,wght@0,400;0,700;1,400&family=Raleway:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
{CSS}
</style>
<script src="https://cdn.onesignal.com/sdks/web/v16/OneSignalSDK.page.js" defer></script>
<script>
  window.OneSignalDeferred = window.OneSignalDeferred || [];
  OneSignalDeferred.push(async function(OneSignal) {{
    await OneSignal.init({{
      appId: "ee15b094-145a-4b1e-9b6c-6a29fa0a469e",
      safari_web_id: "web.onesignal.auto.0868f031-816f-4b4e-9724-08fcd0b320db",
      notifyButton: {{ enable: false }},
    }});
  }});
</script>
<script>
/* Lift the font-loading veil once fonts are ready, or after 800ms — whichever
   comes first. This prevents the flash of fallback text on first visit while
   ensuring the app never stays hidden on a slow connection. */
(function() {{
  var TIMEOUT = 800;
  function reveal() {{
    var app = document.getElementById('app');
    if (app) app.classList.add('fonts-ready');
  }}
  if (document.fonts && document.fonts.ready) {{
    var timer = setTimeout(reveal, TIMEOUT);
    document.fonts.ready.then(function() {{ clearTimeout(timer); reveal(); }});
  }} else {{
    setTimeout(reveal, TIMEOUT);
  }}
}})();
</script>
</head>
<body>
<div class="app" id="app">

  <div class="top-bar">
    <div class="top-bar-inner top-bar-inner--prod">
      <div class="brand-wrap">
        <div class="brand">Vigil</div>
        <div class="day-label" id="dayLabel"></div>
      </div>
    </div>
  </div>

  <div class="screens-container" id="screens">
{all_days_html}
  </div>

  <!-- Threshold veil: Screen 7 → Screen 1 wrap -->
  <div id="thresholdVeil"></div>

  <!-- Navigation dots -->
  <nav class="nav-dots" id="dots" aria-label="Screen navigation"></nav>
  <div class="swipe-hint" id="swipeHint">&#8592; swipe &#8594;</div>
  <div class="screen-counter" id="counter">1 / 7</div>

  <!-- Figure modal -->
  <div class="figure-overlay" id="figureOverlay" role="dialog" aria-modal="true" aria-labelledby="figureName">
    <div class="figure-card">
      <button class="close" id="figureClose" aria-label="Close">&times;</button>
      <div class="figure-eyebrow">People of the Faith</div>
      <div class="figure-name" id="figureName"></div>
      <div class="figure-dates" id="figureDates"></div>
      <div class="figure-role" id="figureRole"></div>
      <div class="figure-divider"></div>
      <div class="figure-body" id="figureBody"></div>
    </div>
  </div>

  <!-- Desktop screen navigation arrows (hidden on mobile via CSS) -->
  <button class="desktop-nav-arrow desktop-nav-prev" id="desktopPrev" aria-label="Previous screen">&#8249;</button>
  <button class="desktop-nav-arrow desktop-nav-next" id="desktopNext" aria-label="Next screen">&#8250;</button>

  <!-- First-visit welcome overlay -->
  <div class="welcome-overlay" id="welcomeOverlay" role="dialog" aria-modal="true" style="display:none">
    <div class="welcome-card">
      <div class="welcome-wordmark">V I G I L</div>
      <div class="welcome-body">Vigil is a short daily devotional shaped by the Christian year. Each day gives you one word to carry through scripture, contemplation, prayer and practice. Move through the seven screens slowly. At the end, mark that you have kept Vigil and return tomorrow.</div>
      <button class="welcome-begin" id="welcomeBegin">Begin</button>
    </div>
  </div>

  <!-- PWA first-launch overlay (shown once when opened from home screen icon) -->
  <div class="welcome-overlay" id="pwaWelcomeOverlay" role="dialog" aria-modal="true" style="display:none">
    <div class="welcome-card">
      <div class="welcome-wordmark">V I G I L</div>
      <div class="welcome-body">You&#8217;re now using Vigil as an app. To receive a word each morning, complete today&#8217;s devotions and &#8216;turn on notifications&#8217; on the final screen.</div>
      <button class="welcome-begin" id="pwaWelcomeBegin">Begin</button>
    </div>
  </div>

  <!-- Install nudge — hidden when running as installed PWA -->
  <div class="install-nudge" id="installNudge" style="display:none">
    <div class="install-nudge-rule"></div>
    <div class="install-nudge-text">Keep Vigil close — <button class="install-nudge-link" id="installNudgeLink">add it to your home screen</button></div>
  </div>

  <!-- Install modal -->
  <div class="install-modal-overlay" id="installModalOverlay">
    <div class="install-modal-card" id="installModalCard">
      <button class="install-modal-close" id="installModalClose" aria-label="Close">&times;</button>
      <div class="install-modal-wordmark">V I G I L</div>
      <div class="install-modal-intro">Adding Vigil to your iPhone home screen is quick and easy. Just follow these steps.</div>
      <div class="install-modal-steps">
        <div class="install-modal-step">
          <div class="install-modal-step-num">1</div>
          <div class="install-modal-step-text">Tap <span class="install-dots-pill">&middot;&middot;&middot;</span> at the bottom right of this screen.</div>
        </div>
        <div class="install-modal-step">
          <div class="install-modal-step-num">2</div>
          <div class="install-modal-step-text">Tap <strong>Share</strong> at the top of the menu.</div>
        </div>
        <div class="install-modal-step">
          <div class="install-modal-step-num">3</div>
          <div class="install-modal-step-text">Scroll down and tap <strong>Add to Home Screen.</strong> (You may need to tap <strong>View More</strong> to see it).</div>
        </div>
        <div class="install-modal-step">
          <div class="install-modal-step-num">4</div>
          <div class="install-modal-step-text">Tap <strong>Add</strong> in the top right corner.</div>
        </div>
      </div>
      <div class="install-modal-rule"></div>
      <div class="install-modal-footnote">Once installed, Vigil will send you a quiet word each morning.</div>
    </div>
  </div>

</div>

<script>
{days_js}
{figures_js}
</script>
<script>
{JS}
</script>
<script>
// ── Service worker registration and update reload ──────────────────────────
// Registers service-worker.js and, when a newly-deployed worker activates,
// reloads the page immediately so users always see today's content rather
// than a cached copy from the previous day.
if ('serviceWorker' in navigator) {{
  navigator.serviceWorker.register('/service-worker.js').then(function(reg) {{
    reg.addEventListener('updatefound', function() {{
      var newWorker = reg.installing;
      newWorker.addEventListener('statechange', function() {{
        // 'activated' fires when the new worker has taken control.
        // Reload so the freshly-cached index.html is served immediately.
        if (newWorker.state === 'activated') {{
          window.location.reload();
        }}
      }});
    }});
  }});
  // If the controller changes (e.g. skipWaiting was called), reload.
  navigator.serviceWorker.addEventListener('controllerchange', function() {{
    window.location.reload();
  }});
}}
</script>
</body>
</html>"""


# ── Holding page ──────────────────────────────────────────────────────────────

def build_holding_page(days, build_time):
    """Build a holding page when no content exists for today."""
    today = today_date()
    first_date = None
    if days:
        first_date = parse_date_from_liturgical_day(days[0]["liturgical_day"])

    if first_date and today < first_date:
        msg = f"Vigil begins on {days[0]['liturgical_day'].split(chr(10))[1].strip() if chr(10) in days[0]['liturgical_day'] else 'soon'}.<br>Come back then."
    else:
        msg = "Vigil will return with a new word tomorrow."

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<!-- Vigil · Holding page · Generated {build_time} -->
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Vigil</title>
<link rel="manifest" href="/manifest.json">
<link rel="apple-touch-icon" href="/icons/icon-180.png">
<link rel="icon" type="image/png" sizes="32x32" href="/icons/icon-32.png">
<link rel="icon" type="image/png" sizes="16x16" href="/icons/icon-16.png">
<meta name="theme-color" content="#FAF7F2">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="default">
<meta name="apple-mobile-web-app-title" content="Vigil">
<link href="https://fonts.googleapis.com/css2?family=Cormorant+Garamond:ital,wght@0,300;0,400;0,500;0,600;1,300;1,400;1,500&family=Raleway:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
:root {{ --bg: #FAF7F2; --gold: #B8860B; --dim: #8A7D6E; }}
html, body {{ margin:0; height:100%; background:var(--bg); font-family:'Raleway',sans-serif; color:var(--dim); }}
.centre {{ display:flex; flex-direction:column; align-items:center; justify-content:center; height:100svh; padding:2rem; text-align:center; }}
.wordmark {{ font-family:'Cormorant Garamond',serif; font-size:2.6rem; font-weight:300; letter-spacing:0.4em; text-transform:uppercase; color:var(--gold); margin-bottom:2rem; }}
.message {{ font-size:1.2rem; line-height:2; max-width:320px; }}
</style>
</head>
<body>
<div class="centre">
  <div class="wordmark">Vigil</div>
  <div class="message">{msg}</div>
</div>
<script>
if ('serviceWorker' in navigator) {{
  navigator.serviceWorker.register('/service-worker.js').then(function(reg) {{
    reg.addEventListener('updatefound', function() {{
      var newWorker = reg.installing;
      newWorker.addEventListener('statechange', function() {{
        if (newWorker.state === 'activated') {{ window.location.reload(); }}
      }});
    }});
  }});
  navigator.serviceWorker.addEventListener('controllerchange', function() {{
    window.location.reload();
  }});
}}
</script>
</body>
</html>"""


# ── Cache version ─────────────────────────────────────────────────────────────

SERVICE_WORKER_FILE = "service-worker.js"

def increment_cache_version():
    """Read service-worker.js, increment the CACHE_VERSION number by 1,
    write the file back, and return (old_version, new_version).
    If the file cannot be found or parsed, prints a warning and returns
    (None, None) so the build can continue without blocking."""
    if not os.path.exists(SERVICE_WORKER_FILE):
        print(f"Warning: {SERVICE_WORKER_FILE} not found — cache version not incremented.")
        return None, None

    with open(SERVICE_WORKER_FILE, "r", encoding="utf-8") as f:
        content = f.read()

    pattern = r"(const CACHE_VERSION\s*=\s*'vigil-v)(\d+)(')"
    m = re.search(pattern, content)
    if not m:
        print(f"Warning: Could not find CACHE_VERSION in {SERVICE_WORKER_FILE} — not incremented.")
        return None, None

    old_n = int(m.group(2))
    new_n = old_n + 1
    old_ver = f"vigil-v{old_n}"
    new_ver = f"vigil-v{new_n}"

    new_content = re.sub(pattern, lambda _: f"{m.group(1)}{new_n}{m.group(3)}", content)

    with open(SERVICE_WORKER_FILE, "w", encoding="utf-8") as f:
        f.write(new_content)

    return old_ver, new_ver


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    build_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    print(f"Reading {SPREADSHEET}...")
    days = load_spreadsheet(SPREADSHEET)

    if not days:
        print("Error: No content rows found in the spreadsheet.")
        sys.exit(1)

    print(f"Building {len(days)} day(s):")

    today_idx = find_today(days)

    for i, day in enumerate(days):
        d = parse_date_from_liturgical_day(day["liturgical_day"])
        date_str = d.strftime("%-d %B %Y") if d else "?"
        raw_word = day["word"].strip().split("\n")[0].strip()
        m = re.match(r'^([A-Z\s\-\']+?)\s*(\[|$)', raw_word)
        word_display = m.group(1).strip() if m else raw_word
        mark = "✓" if i == today_idx else "◦"
        print(f"  {mark} Day {i+1:2d} · {date_str} · {word_display}")

    # --all-days embeds every content row so ?date= can reach past and future days.
    all_days_flag = "--all-days" in sys.argv

    if today_idx == -1:
        print("\n  ◦ No content for today — generating holding page.")
        html_out = build_holding_page(days, build_time)
    else:
        if all_days_flag:
            # Embed all rows from the spreadsheet, not just the rolling window.
            # Today is still the default view; ?date= can reach any embedded day.
            window_days = days
            print(f"\n  --all-days: embedding all {len(days)} day(s).")
        else:
            # Standard production build: rolling window only.
            window_days = days[today_idx : today_idx + DAYS_WINDOW]
        days_window = [(day, parse_hover_links(day["hover_links"])) for day in window_days]
        today_window_idx = today_idx if all_days_flag else 0
        html_out = build_html(days_window, window_days[today_window_idx], build_time)

    # Write output
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(html_out)

    size_kb = os.path.getsize(OUTPUT_FILE) / 1024

    days_embedded = len(days_window) if today_idx != -1 else 0
    print(f"\n  ✓ Wrote {OUTPUT_FILE} ({size_kb:.1f} KB) — {days_embedded} day(s) embedded")

    # Increment cache version automatically on every build
    old_ver, new_ver = increment_cache_version()
    if new_ver:
        print(f"  ✓ Cache version: {old_ver} → {new_ver}")
    print()
    print(f"Deploy to github.com/This-Is-Pelagius/Vigil (v2-build branch):")
    print(f"  dailyvigil.app  ← {OUTPUT_FILE}")
    if new_ver:
        print(f"  dailyvigil.app  ← {SERVICE_WORKER_FILE}")
        print()
        print(f"Commit message:")
        today_ref = window_days[today_window_idx] if today_idx != -1 else None
        today_d = parse_date_from_liturgical_day(today_ref["liturgical_day"]) if today_ref else None
        commit_date = today_d.strftime("%-d %b %Y") if today_d else "today"
        raw_word = today_ref["word"].strip().split("\n")[0].strip() if today_ref else "WORD"
        wm = re.match(r'^([A-Z\s\-\']+?)\s*(\[|$)', raw_word)
        commit_word = wm.group(1).strip() if wm else raw_word
        print(f'  git add index.html service-worker.js')
        print(f'  git commit -m "Deploy: {commit_date} · {commit_word} · cache {new_ver}"')
        print(f'  git push origin v2-build')
        print()
        print(f'Then schedule notifications:')
        print(f'  ~/vigil-env/bin/python3 vigil_notify.py 14')
    print()


if __name__ == "__main__":
    main()
