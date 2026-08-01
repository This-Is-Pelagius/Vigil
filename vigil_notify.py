#!/usr/bin/env python3
"""
vigil_notify.py
Vigil notification scheduler — Version 2.0

Reads Vigil_Content_v2.0.xlsx and schedules Word of the Day push notifications
via the OneSignal API.

Default behaviour: schedules today's notification only. Designed to be run
automatically every morning via GitHub Actions (see .github/workflows/vigil-notify.yml).

Usage:
    python3 vigil_notify.py           # Schedule today only (default)
    python3 vigil_notify.py 7         # Schedule the next 7 days
    python3 vigil_notify.py --dry-run # Preview what would be scheduled, no API calls

Requirements:
    pip3 install openpyxl requests

Configuration:
    When running locally: create a file called vigil_notify_config.py in the
    same folder as this script containing your OneSignal REST API key:

        ONESIGNAL_API_KEY = "your-rest-api-key-here"

    Do NOT commit vigil_notify_config.py to GitHub. Add it to your .gitignore.

    When running via GitHub Actions: the API key is read from the repository
    secret ONESIGNAL_API_KEY. No config file is needed.

BST/GMT:
    The script automatically detects whether London is currently on British
    Summer Time (UTC+1) or GMT (UTC+0) and adjusts delivery timing accordingly.
    No manual changes are ever needed.
"""

import re
import sys
import datetime
import os

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Run: pip3 install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("Error: requests is required. Run: pip3 install requests")
    sys.exit(1)

try:
    from zoneinfo import ZoneInfo
    _LONDON_TZ = ZoneInfo("Europe/London")
except ImportError:
    _LONDON_TZ = None

# ── Configuration ─────────────────────────────────────────────────────────────

SPREADSHEET       = "Vigil_Content_v2.0.xlsx"
ONESIGNAL_APP_ID  = "ee15b094-145a-4b1e-9b6c-6a29fa0a469e"
ONESIGNAL_API_URL = "https://onesignal.com/api/v1/notifications"

# Delivery time: 7:00 AM in each user's local timezone.
DELIVERY_HOUR   = 7
DELIVERY_MINUTE = 0

# How many hours before the delivery time the API call must be made.
BUFFER_HOURS = 1

# Default: schedule today only. Pass a number on the command line to schedule
# further ahead (e.g. python3 vigil_notify.py 7).
DEFAULT_DAYS = 1

# ── Load API key ──────────────────────────────────────────────────────────────

def load_api_key():
    """
    Load the OneSignal REST API key.
    Checks the environment variable first (for GitHub Actions), then falls
    back to vigil_notify_config.py (for local use).
    """
    # GitHub Actions sets secrets as environment variables.
    key = os.environ.get("ONESIGNAL_API_KEY")
    if key:
        return key

    # Local use: read from vigil_notify_config.py.
    try:
        from vigil_notify_config import ONESIGNAL_API_KEY
        return ONESIGNAL_API_KEY
    except ImportError:
        pass

    print("Error: No OneSignal API key found.")
    print()
    print("For local use: create vigil_notify_config.py in your Vigil folder")
    print("containing one line:")
    print()
    print('    ONESIGNAL_API_KEY = "your-rest-api-key-here"')
    print()
    print("For GitHub Actions: add ONESIGNAL_API_KEY as a repository secret")
    print("under Settings → Secrets and variables → Actions.")
    print()
    print("Your REST API key is in the OneSignal dashboard under")
    print("Settings → Keys & IDs → API Keys.")
    sys.exit(1)


# ── BST/GMT detection ─────────────────────────────────────────────────────────

def london_utc_offset_hours(date):
    """
    Returns the UTC offset for London on the given date: 1 during BST, 0 during GMT.
    Uses Python's zoneinfo module (Python 3.9+) for accurate DST detection.
    Falls back to a manual approximation if zoneinfo is unavailable.
    """
    if _LONDON_TZ is not None:
        # zoneinfo gives the definitive answer.
        dt = datetime.datetime(date.year, date.month, date.day, 12, 0, tzinfo=_LONDON_TZ)
        offset = dt.utcoffset()
        return int(offset.total_seconds() / 3600)

    # Fallback: approximate BST as last Sunday in March to last Sunday in October.
    # This matches the UK rule precisely for all years since 2002.
    def last_sunday(year, month):
        # Find the last day of the month, then back up to Sunday.
        import calendar
        last_day = calendar.monthrange(year, month)[1]
        d = datetime.date(year, month, last_day)
        # weekday(): Monday=0, Sunday=6
        d -= datetime.timedelta(days=(d.weekday() + 1) % 7)
        return d

    bst_start = last_sunday(date.year, 3)   # Last Sunday in March
    bst_end   = last_sunday(date.year, 10)  # Last Sunday in October

    if bst_start <= date < bst_end:
        return 1
    return 0


# ── Spreadsheet parsing ───────────────────────────────────────────────────────

MONTH_NAMES = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}


def load_spreadsheet(path):
    """Load all data rows from all tabs. Returns list of day dicts."""
    if not os.path.exists(path):
        print(f"Error: Cannot find spreadsheet '{path}'.")
        print("Make sure Vigil_Content_v2.0.xlsx is in the same folder as this script.")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    days = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            continue

        header = [str(c).strip() if c else "" for c in rows[0]]
        col = {}

        for i, h in enumerate(header):
            hl = h.lower()
            if "liturgical day" in hl:
                col["liturgical_day"] = i
            elif "screen 2" in hl or "word" in hl:
                col["word"] = i
            elif "screen 3" in hl or "scripture" in hl:
                col["scripture"] = i

        required = ["liturgical_day", "word", "scripture"]
        for r in required:
            if r not in col:
                col = {}
                break
        if not col:
            continue

        for row in rows[1:]:
            if all(v is None or str(v).strip() == "" for v in row):
                continue

            def get(key, default=""):
                if key not in col:
                    return default
                v = row[col[key]]
                if v is None:
                    return default
                s = str(v).strip()
                if s.startswith("'"):
                    s = s[1:]
                return s

            liturgical_day = get("liturgical_day")
            word_raw       = get("word")
            scripture      = get("scripture")

            if not liturgical_day or not word_raw:
                continue

            days.append({
                "liturgical_day": liturgical_day,
                "word_raw":       word_raw,
                "scripture":      scripture,
            })

    return days


def parse_date(liturgical_day_text):
    """Extract a date from the liturgical day cell. Returns datetime.date or None."""
    m = re.search(r'\b(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})\b', liturgical_day_text)
    if not m:
        return None
    day_n   = int(m.group(1))
    month_n = MONTH_NAMES.get(m.group(2).lower())
    year_n  = int(m.group(3))
    if not month_n:
        return None
    try:
        return datetime.date(year_n, month_n, day_n)
    except ValueError:
        return None


def parse_word(word_raw):
    """Extract the display word (uppercase letters before any bracket or newline)."""
    first_line = word_raw.strip().split("\n")[0].strip()
    m = re.match(r'^([A-Z\s\-\']+?)\s*(\[|$)', first_line)
    return m.group(1).strip() if m else first_line


def parse_scripture(scripture_raw):
    """
    Extract the scripture reference from the last non-empty line of the cell.
    The cell contains the full passage text followed by the reference on the
    last line, e.g. 'John 15:4–5 (NRSV)' or 'Mark 16:15, 20 — NRSV'.
    Returns only the reference, stripping the translation label.
    """
    lines = [l.strip() for l in scripture_raw.strip().split("\n") if l.strip()]
    if not lines:
        return ""
    ref = lines[-1]
    ref = re.sub(r'\s*[\u2014\-]\s*(NRSV|ESV|NIV|KJV|NLT|NASB|CSB)\s*$', '', ref)
    ref = re.sub(r'\s*\((NRSV|ESV|NIV|KJV|NLT|NASB|CSB)\)\s*$', '', ref)
    return ref.strip()


# ── Notification format ────────────────────────────────────────────────────────

def format_message(word, scripture):
    """Build the notification body: WORD · Book Ch:v"""
    return f"{word} · {scripture}"


def format_message_name(date, word):
    """Build the internal message name: DD Mon YYYY — WORD"""
    return f"{date.strftime('%-d %b %Y')} — {word}"


# ── OneSignal delivery time ────────────────────────────────────────────────────

def delivery_send_after(date):
    """
    Returns the OneSignal 'send_after' timestamp for the given date.

    We want OneSignal to begin delivery at (DELIVERY_HOUR - BUFFER_HOURS) in
    London local time, which gives it a one-hour head start before 7:00 AM
    arrives in the earliest timezone where Vigil has users.

    The UTC offset is detected automatically from the date: +1 in BST, +0 in GMT.
    """
    utc_offset = london_utc_offset_hours(date)
    send_hour_utc = (DELIVERY_HOUR - BUFFER_HOURS) - utc_offset

    send_dt = datetime.datetime(
        date.year, date.month, date.day,
        send_hour_utc, 0, 0,
    )
    # OneSignal format: "2026-05-01 05:00:00 GMT+0000"
    return send_dt.strftime("%Y-%m-%d %H:%M:%S GMT+0000")


# ── Already-passed check ──────────────────────────────────────────────────────

def is_too_late(date):
    """
    Returns True if it is too late to schedule today's notification.
    'Too late' means the current local time has passed the buffer cutoff.
    """
    now = datetime.datetime.now()
    cutoff = datetime.datetime(
        date.year, date.month, date.day,
        DELIVERY_HOUR - BUFFER_HOURS, DELIVERY_MINUTE, 0,
    )
    return date == datetime.date.today() and now >= cutoff


# ── OneSignal API call ─────────────────────────────────────────────────────────

def schedule_notification(date, word, scripture, api_key, dry_run=False):
    """Schedule a single notification via the OneSignal API."""
    message_name = format_message_name(date, word)
    message_body = format_message(word, parse_scripture(scripture))
    send_after   = delivery_send_after(date)

    utc_offset   = london_utc_offset_hours(date)
    tz_label     = "BST (UTC+1)" if utc_offset == 1 else "GMT (UTC+0)"

    payload = {
        "app_id":               ONESIGNAL_APP_ID,
        "name":                 message_name,
        "headings":             {"en": "Word of the Day"},
        "contents":             {"en": message_body},
        "url":                  "https://dailyvigil.app",
        "included_segments":    ["All"],
        "send_after":           send_after,
        "delayed_option":       "timezone",
        "delivery_time_of_day": f"{DELIVERY_HOUR:02d}:{DELIVERY_MINUTE:02d}",
        "collapse_id":          f"vigil-daily-{date.isoformat()}",
    }

    date_display = date.strftime("%-d %b %Y")

    if dry_run:
        print(f"  [DRY RUN] {date_display} · {message_body}")
        print(f"            Send after: {send_after}  ({tz_label})")
        print(f"            Deliver at: {DELIVERY_HOUR:02d}:{DELIVERY_MINUTE:02d} per user timezone")
        return True

    headers = {
        "Content-Type":  "application/json",
        "Authorization": f"Key {api_key}",
    }

    try:
        response = requests.post(ONESIGNAL_API_URL, json=payload, headers=headers, timeout=15)
        data = response.json()

        if response.status_code == 200 and "id" in data:
            notif_id = data["id"]
            print(f"  ✓ {date_display} · {message_body}")
            print(f"    Notification ID: {notif_id}  ({tz_label})")
            return True
        else:
            errors = data.get("errors", [data])
            print(f"  ✗ {date_display} · {message_body}")
            print(f"    Error: {errors}")
            return False

    except requests.RequestException as e:
        print(f"  ✗ {date_display} · Network error: {e}")
        return False


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    dry_run  = "--dry-run" in sys.argv
    day_args = [a for a in sys.argv[1:] if a != "--dry-run"]
    num_days = int(day_args[0]) if day_args else DEFAULT_DAYS

    mode_label = " (DRY RUN — no notifications will be sent)" if dry_run else ""
    print(f"Vigil Notification Scheduler{mode_label}")
    if num_days == 1:
        print("Scheduling today's notification")
    else:
        print(f"Scheduling {num_days} day(s) of notifications")
    print()

    api_key = load_api_key()

    print(f"Reading {SPREADSHEET}...")
    days = load_spreadsheet(SPREADSHEET)

    if not days:
        print("Error: No content rows found in the spreadsheet.")
        sys.exit(1)

    today = datetime.date.today()

    # Find today's index in the spreadsheet.
    today_idx = -1
    for i, day in enumerate(days):
        d = parse_date(day["liturgical_day"])
        if d and d == today:
            today_idx = i
            break

    if today_idx == -1:
        print(f"Warning: No content found for today ({today}). Starting from the next available day.")
        for i, day in enumerate(days):
            d = parse_date(day["liturgical_day"])
            if d and d > today:
                today_idx = i
                break

    if today_idx == -1:
        print("Error: No upcoming content found in the spreadsheet.")
        sys.exit(1)

    window = days[today_idx : today_idx + num_days]

    if not window:
        print("Error: No days found to schedule.")
        sys.exit(1)

    print(f"Found {len(window)} day(s) to schedule:\n")

    scheduled = 0
    skipped   = 0
    failed    = 0

    for day in window:
        date      = parse_date(day["liturgical_day"])
        word      = parse_word(day["word_raw"])
        scripture = parse_scripture(day["scripture"])

        if not date:
            print(f"  ⚠ Could not parse date from: {day['liturgical_day'][:40]} — skipping")
            skipped += 1
            continue

        if not scripture:
            print(f"  ⚠ {date.strftime('%-d %b %Y')} · {word} — no scripture reference found, skipping")
            skipped += 1
            continue

        if is_too_late(date):
            print(f"  ⚠ {date.strftime('%-d %b %Y')} · {word} — too late to schedule today's notification (past {DELIVERY_HOUR - BUFFER_HOURS}:00), skipping")
            skipped += 1
            continue

        success = schedule_notification(date, word, scripture, api_key, dry_run=dry_run)
        if success:
            scheduled += 1
        else:
            failed += 1

    print()
    print("─" * 48)
    if dry_run:
        print(f"Dry run complete. {scheduled} notification(s) would be scheduled.")
    else:
        print(f"Done. {scheduled} scheduled · {skipped} skipped · {failed} failed.")
    if failed:
        print("Check the errors above and retry failed days manually in the OneSignal dashboard.")
    print()


if __name__ == "__main__":
    main()
